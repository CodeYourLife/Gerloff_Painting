from changeorder.models import ClientJobRoles, ChangeOrders
from console.models import *
from console.misc import Email

from datetime import date
from dateutil.parser import parse as parse_date

from decimal import Decimal

from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.core.serializers.json import DjangoJSONEncoder
from django.db.models import Q
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_datetime
from django.conf import settings
from django.utils import timezone

from django_tables2 import RequestConfig

from employees.models import *
from employees.models import Employees
from equipment.models import Inventory
from equipment.tables import JobsTable
from equipment.filters import JobsFilter
from jobs.models import *
from jobs.models import ClockSharkTimeEntry, Jobs
from jobs.JobMisc import start_date_change, gerloff_super_change
from jobs.filters import JobNotesFilter
from rentals.models import Rentals
from subcontractors.models import *
from submittals.models import *
from wallcovering.models import Wallcovering, Packages, OutgoingItem, OrderItems

import csv
import json
import os
import os.path
import openpyxl
import requests





@login_required(login_url='/accounts/login')
def change_start_date(request, jobnumber, previous, super, filter):
    # jobnumber is the job number you are changing
    # previous is either "jobpage"(job_page) or "super"(super_home)
    # super is if super_home was filtered to a certain superintendent
    # filter is if super_home was filtered to UPCOMING or ALL. OR, it is set to JOB for job page

    jobs = Jobs.objects.get(job_number=jobnumber)
    format_date = jobs.start_date.strftime("%Y-%m-%d")
    previous_page = previous
    if request.method == 'POST':
        if 'is_active' in request.POST:
            if jobs.is_active == False:
                status = 1
            else:
                status = 3
        elif jobs.is_active == True:
            status = 2
        else:
            status = 3
        if str(jobs.start_date) != str(request.POST['start_date']):
            start_date_change(jobs, request.POST['start_date'], status, request.POST['date_note'],
                              Employees.objects.get(user=request.user), True)
        elif status != 3:
            start_date_change(jobs, request.POST['start_date'], status, request.POST['date_note'],
                              Employees.objects.get(user=request.user), False)
        else:
            if 'follow_up' in request.POST or request.POST['date_note'] != "":
                jobs.start_date_checked = date.today()
                if request.POST['date_note'] != "":
                    JobNotes.objects.create(job_number=jobs,
                                            note="Start Date is Still: " + str(request.POST['start_date']) + ". " +
                                                 request.POST['date_note'],
                                            type="auto_start_date_note",
                                            user=Employees.objects.get(user=request.user),
                                            date=date.today())
                jobs.save()
        if previous == 'jobpage':
            if filter == 'JOB':
                return redirect('job_page', jobnumber=jobnumber)
            else:
                return redirect('jobs_home')
        else:
            return redirect('super_home', super=super)
    return render(request, "change_start_date.html",
                  {'jobs': jobs, 'formatdate': format_date,
                   'notes': JobNotes.objects.filter(job_number=jobnumber, type="auto_start_date_note"),
                   'previous_page': previous_page, 'selected_super': super,
                   'selected_filter': filter})


def change_gpsuper(request, jobnumber, previous):
    jobs = Jobs.objects.get(job_number=jobnumber)
    previous_page = previous
    employees = Employees.objects.exclude(job_title__description="Painter")
    if request.method == 'POST':
        gerloff_super_change(jobs, Employees.objects.get(id=request.POST['select_gpsuper']),
                             Employees.objects.get(user=request.user))
        if previous == 'jobpage':
            return redirect('job_page', jobnumber=jobnumber)
        else:
            return redirect('super_home', super='ALL')
    return render(request, "change_gpsuper.html",
                  {'jobs': jobs, 'previous_page': previous_page, 'employees': employees})


@login_required(login_url='/accounts/login')
def update_job_info(request, jobnumber):
    send_data = {}
    selectedjob = Jobs.objects.get(job_number=jobnumber)
    send_data['selectedjob'] = selectedjob
    allclients = Clients.objects.order_by('company')
    send_data['allclients'] = allclients
    pms = ClientEmployees.objects.values('name', 'id', 'person_pk')
    estimators = Employees.objects.exclude(job_title__description='Painter')
    send_data['estimators'] = estimators
    superintendents = Employees.objects.exclude(job_title__description='Painter')
    send_data['superintendents'] = superintendents
    notes = JobNotes.objects.filter(job_number=jobnumber)
    # send_employees = Employees.objects.filter(job_title="Superintendent")[0:2000]
    send_data['notes'] = notes
    prices_json = json.dumps(list(pms), cls=DjangoJSONEncoder)
    send_data['data'] = prices_json
    selectedclient = Clients.objects.get(id=selectedjob.client.id)
    pms_filter = ClientEmployees.objects.filter(id=selectedclient.id)
    send_data['pms_filter'] = pms_filter
    startdate = selectedjob.start_date.strftime("%Y") + "-" + selectedjob.start_date.strftime(
        "%m") + "-" + selectedjob.start_date.strftime("%d")
    send_data['startdate'] = startdate
    if request.method == 'POST':
        if selectedjob.job_name != request.POST['job_name']:
            selectedjob.job_name = request.POST['job_name']
        if selectedjob.address != request.POST['address']:
            selectedjob.address = request.POST['address']
        if selectedjob.city != request.POST['city']:
            selectedjob.city = request.POST['city']
        if selectedjob.state != request.POST['state']:
            selectedjob.state = request.POST['state']
        if 'off_hours' in request.POST:
            selectedjob.is_off_hours = True
        else:
            selectedjob.is_off_hours = False
        if 'on_base2' in request.POST:
            selectedjob.is_on_base = True
        else:
            selectedjob.is_on_base = False
        if 'is_painting_subbed' in request.POST:
            selectedjob.is_painting_subbed = True
        else:
            selectedjob.is_painting_subbed = False
        if 'is_wage_rate' in request.POST:
            selectedjob.is_wage_scale = True
        else:
            selectedjob.is_wage_scale = False
        if request.POST['brush_role'] == "":
            if selectedjob.brush_role is not None:
                selectedjob.brush_role = None
        else:
            if selectedjob.brush_role != request.POST['brush_role']:
                selectedjob.brush_role = request.POST['brush_role']
        if request.POST['spray_scale'] == "":
            if selectedjob.spray_scale is not None:
                selectedjob.spray_scale = None
        else:
            if selectedjob.spray_scale != request.POST['spray_scale']:
                selectedjob.spray_scale = request.POST['spray_scale']
        if 'is_closed' in request.POST:
            if selectedjob.is_closed == False:
                if Inventory.objects.filter(job_number=selectedjob, is_closed=False).exists():
                    message = "Job: " + selectedjob.job_name + " is closed. The following equipment is assigned to the job and must be returned immediately!\n "
                    recipients = ["admin1@gerloffpainting.com", "admin2@gerloffpainting.com",
                                  "warehouse@gerloffpainting.com", "joe@gerloffpainting.com"]
                    if selectedjob.superintendent.email is None:
                        message = message + "\n No email address for " + str(selectedjob.superintendent)
                    else:
                        recipients.append(selectedjob.superintendent.email)
                    for x in Inventory.objects.filter(job_number=selectedjob, is_closed=False):
                        if x.number:
                            message = message + "\n -" + x.item + " GP Number #" + x.number
                        else:
                            message = message + "\n -" + x.item + " -No GP Number! "
                    try:
                        Email.sendEmail("Closed Job - " + selectedjob.job_name, message,
                                        recipients, False)
                        success = True
                    except:
                        send_data['error_message'] = "Email Failed to send. Please let warehouse know. " + message
                if Subcontracts.objects.filter(is_closed=False, job_number=selectedjob).exists():
                    message = "Job: " + selectedjob.job_name + " cannot be closed. The following subcontracts are still open!\n "
                    recipients = ["admin1@gerloffpainting.com", "admin2@gerloffpainting.com",
                                  "joe@gerloffpainting.com"]
                    for x in Subcontracts.objects.filter(is_closed=False, job_number=selectedjob):
                        if x.po_number:
                            message += "\n -" + x.subcontractor.company + " - PO# " + x.po_number + "! "
                        else:
                            message += "\n -" + x.subcontractor.company + " - PO# N/A!"
                    try:
                        Email.sendEmail("Closed Job Error- " + selectedjob.job_name, message,
                                        recipients, False)
                        success = True
                    except:
                        send_data['error_message'] = "Email Failed to send. Please note that the job can't be closed because there are open subcontractors"
                    send_data['subcontract_open_error'] = True
                    send_data['open_subcontracts'] = message
                else:
                    selectedjob.is_closed = True
        else:
            selectedjob.is_closed = False
        if 'is_t_m_job' in request.POST:
            selectedjob.is_t_m_job = True
        else:
            selectedjob.is_t_m_job = False
        if request.POST['t_m_nte_amount'] == "":
            if selectedjob.t_m_nte_amount is not None:
                selectedjob.t_m_nte_amount = None
        else:
            if selectedjob.t_m_nte_amount != request.POST['t_m_nte_amount']:
                selectedjob.t_m_nte_amount = request.POST['t_m_nte_amount']
        if selectedjob.contract_status != request.POST['contract_status']:
            selectedjob.contract_status = request.POST['contract_status']
        if selectedjob.insurance_status != request.POST['insurance_status']:
            selectedjob.insurance_status = request.POST['insurance_status']
        if 'has_submittals' in request.POST:
            selectedjob.submittals_needed = True
        else:
            selectedjob.submittals_needed = False
        if 'is_bonded' in request.POST:
            selectedjob.is_bonded = True
        else:
            selectedjob.is_bonded = False
        if request.POST['select_company'] == 'add_new':
            client = Clients.objects.create(company=request.POST['new_client'],
                                            bid_email=request.POST['new_client_bid_email'],
                                            phone=request.POST['new_client_phone'])
            selectedjob.client = client
            selectedjob.save()
        elif selectedjob.client.id != request.POST['select_company']:
            selectedjob.client = Clients.objects.get(id=request.POST['select_company'])
            selectedjob.save()
        selectedjob.client.company = request.POST['new_client']
        selectedjob.client.bid_email = request.POST['new_client_bid_email']
        selectedjob.client.phone = request.POST['new_client_phone']
        selectedjob.client.save()
        if request.POST['select_pm'] == 'add_new':
            client_pm = ClientEmployees.objects.create(id=selectedjob.client, name=request.POST['new_pm'],
                                                       phone=request.POST['new_pm_phone'],
                                                       email=request.POST['new_pm_email'])
            selectedjob.client_Pm = client_pm
        else:
            if selectedjob.client_Pm.person_pk != request.POST['select_pm']:
                selectedjob.client_Pm = ClientEmployees.objects.get(person_pk=request.POST['select_pm'])
        selectedjob.client_Pm.name = request.POST['new_pm']
        selectedjob.client_Pm.phone = request.POST['new_pm_phone']
        selectedjob.client_Pm.email = request.POST['new_pm_email']
        selectedjob.client_Pm.save()
        if request.POST['select_super'] == 'add_new':
            client_super = ClientEmployees.objects.create(id=selectedjob.client, name=request.POST['new_super'],
                                                          phone=request.POST['new_super_phone'],
                                                          email=request.POST['new_super_email'])
            selectedjob.client_Super = client_super
        elif request.POST['select_super'] == 'not_sure':
            if selectedjob.client_Super is not None:
                selectedjob.client_Super = None
        else:
            if selectedjob.client_Super is not None:
                if selectedjob.client_Super.person_pk != request.POST['select_super']:
                    selectedjob.client_Super = ClientEmployees.objects.get(person_pk=request.POST['select_super'])
            else:
                selectedjob.client_Super = ClientEmployees.objects.get(person_pk=request.POST['select_super'])
            selectedjob.client_Super.name = request.POST['new_super']
            selectedjob.client_Super.phone = request.POST['new_super_phone']
            selectedjob.client_Super.email = request.POST['new_super_email']
            selectedjob.client_Super.save()
        if request.POST['select_gpsuper'] == 'not_sure':
            if selectedjob.superintendent is not None:
                selectedjob.superintendent = None
        else:
            if selectedjob.superintendent is not None:
                if str(selectedjob.superintendent.id) != str(request.POST['select_gpsuper']):
                    email_sent = gerloff_super_change(selectedjob,
                                                      Employees.objects.get(id=request.POST['select_gpsuper']),
                                                      Employees.objects.get(user=request.user))
                    if email_sent:
                        send_data['email_sent'] = email_sent
                    else:
                        send_data['email_not_sent'] = email_sent
            else:
                email_sent = gerloff_super_change(selectedjob, Employees.objects.get(id=request.POST['select_gpsuper']),
                                                  Employees.objects.get(user=request.user))
                if email_sent:
                    send_data['email_sent'] = email_sent
                else:
                    send_data['email_not_sent'] = email_sent
        if selectedjob.estimator.id != request.POST['select_gpestimator']:
            selectedjob.estimator = Employees.objects.get(id=request.POST['select_gpestimator'])
        if 'has_paint' in request.POST:
            selectedjob.has_paint = True
        else:
            selectedjob.has_paint = False
        if 'has_wallcovering' in request.POST:
            selectedjob.has_wallcovering = True
        else:
            selectedjob.has_wallcovering = False
        if 'has_special_paint' in request.POST:
            selectedjob.special_paint_needed = True
        else:
            selectedjob.special_paint_needed = False

        if startdate != request.POST['start_date']:
            email_sent = start_date_change(selectedjob, request.POST['start_date'], 3, request.POST['date_note'],
                                           Employees.objects.get(user=request.user), True, True)
            if email_sent:
                send_data['email_sent'] = email_sent
            else:
                send_data['email_not_sent'] = email_sent
        if selectedjob.notes != request.POST['email_job_note']:
            selectedjob.notes = request.POST['email_job_note']
        if request.POST['po_number'] == "":
            if selectedjob.po_number is not None:
                selectedjob.po_number = None
        else:
            if selectedjob.po_number != request.POST['po_number']:
                selectedjob.po_number = request.POST['po_number']
        if request.POST['contract_amount'] == "":
            if selectedjob.contract_amount is not None:
                selectedjob.contract_amount = None
        else:
            if selectedjob.contract_amount != request.POST['contract_amount']:
                selectedjob.contract_amount = request.POST['contract_amount']
        if request.POST['painting_budget'] == "":
            if selectedjob.painting_budget is not None:
                selectedjob.painting_budget = None
        else:
            if selectedjob.painting_budget != request.POST['painting_budget']:
                selectedjob.painting_budget = request.POST['painting_budget']

        if request.POST['wallcovering_budget'] == "":
            if selectedjob.wallcovering_budget is not None:
                selectedjob.wallcovering_budget = None
        else:
            if selectedjob.wallcovering_budget != request.POST['wallcovering_budget']:
                selectedjob.wallcovering_budget = request.POST['wallcovering_budget']
        selectedjob.save()
    return render(request, 'update_job_info.html', send_data)


def audit_MC_open_jobs2(request):
    send_data = {}
    if request.method == 'POST':
        if 'upload_file' in request.FILES:
            fileitem = request.FILES['upload_file']
            fn2 = os.path.join(settings.MEDIA_ROOT, "job_upload", "Temp.xlsx")
            open(fn2, 'wb').write(fileitem.file.read())
            wb_obj = openpyxl.load_workbook(filename=request.FILES['upload_file'].file)
            sheet_obj = wb_obj["Booked"]
            # client_name = sheet_obj.cell(row=16, column=2).value
            needs_to_be_opened = []
            needs_to_be_closed = []
            needs_to_be_labor_done = []
            not_found = []
            open_jobs = []
            labor_done = []
            needs_to_not_be_labor_done = []
            closed_but_equipment = []
            closed_but_subs = []
            superintendents = []
            superintendenterrors=[]
            error_message = ""
            a = 3
            while sheet_obj.cell(row=a, column=1).value != None:
                mc_labor_done = 0
                superintendents.append({'job_number': sheet_obj.cell(row=a, column=1).value,
                                        'super': sheet_obj.cell(row=a, column=7).value})
                a = a + 1
                if a > 200000:
                    break
            for x in Jobs.objects.filter(is_closed=False):  # scroll through open jobs in Trinity
                if x.superintendent:

                    print("NOTHING")
                    # trinitysuper = x.superintendent.first_name
                    # mcsuper = "couldn't find"
                    # for y in superintendents:
                    #     if y['job_number'] == x.job_number:
                    #         mcsuper = y['super']
                    #         if mcsuper == 'Sub' and not trinitysuper == 'Victor':
                    #             superintendenterrors.append('Job ' + str(x.job_number) + ' has ' + mcsuper + ' in MC, but ' + trinitysuper + ' in Trinity')
                    #         elif mcsuper == 'Steve' and not trinitysuper == 'Steve':
                    #             superintendenterrors.append('Job ' + str(x.job_number) + ' has ' + mcsuper + ' in MC, but ' + trinitysuper + ' in Trinity')
                    #         elif mcsuper == 'Ed' and not trinitysuper == 'Edward':
                    #             superintendenterrors.append('Job ' + str(x.job_number) + ' has ' + mcsuper + ' in MC, but ' + trinitysuper + ' in Trinity')
                    #         else: superintendenterrors.append('CONGRATS! Job ' + str(x.job_number) + ' has ' + mcsuper + ' in MC, and ' + trinitysuper + ' in Trinity')
                    #         break
                else:
                    for y in superintendents:
                        if y['job_number'] == x.job_number:
                            if y['super']:
                                superintendenterrors.append('Job ' + str(x.job_number) + ' has ' + y['super'] + ' in MC, but no Super in Trinity')
                            else:
                                superintendenterrors.append('CONGRATS AGAIN! Job ' + str(
                                    x.job_number) + ' has no Super in MC, and also no Super in Trinity')

                            break
            print(superintendenterrors)
    return redirect("/")

@login_required(login_url='/accounts/login')
def audit_MC_open_jobs(request):
    send_data = {}
    if request.method == 'POST':
        if 'upload_file' in request.FILES:
            fileitem = request.FILES['upload_file']
            fn2 = os.path.join(settings.MEDIA_ROOT, "job_upload", "Temp.xlsx")
            open(fn2, 'wb').write(fileitem.file.read())
            wb_obj = openpyxl.load_workbook(filename=request.FILES['upload_file'].file)
            sheet_obj = wb_obj["Booked"]
            # client_name = sheet_obj.cell(row=16, column=2).value
            needs_to_be_opened = []
            needs_to_be_closed = []
            needs_to_be_labor_done = []
            not_found = []
            open_jobs = []
            labor_done = []
            needs_to_not_be_labor_done = []
            closed_but_equipment = []
            closed_but_subs = []
            superintendents = []
            superintendenterrors = []
            error_message = ""
            a = 3
            while sheet_obj.cell(row=a, column=1).value != None:
                mc_labor_done = 0
                superintendents.append({'job_number':sheet_obj.cell(row=a, column=1).value, 'super':sheet_obj.cell(row=a, column=7).value})
                if sheet_obj.cell(row=a, column=18).value == "Open":  #if open in MC
                    job_number = sheet_obj.cell(row=a, column=1).value
                    open_jobs.append(job_number)
                    if sheet_obj.cell(row=a, column=25).value != None:  #job has been marked labor done in MC
                        mc_labor_done = 1
                    if Jobs.objects.filter(job_number=job_number).exists():  #job exists in Trinity
                        job = Jobs.objects.get(job_number=job_number)
                        if job.is_closed == True:
                            job.is_closed = False
                            job.save()
                            needs_to_be_opened.append({'job_number': job_number, 'job_name': job.job_name})
                        if mc_labor_done == 1:
                            labor_done.append({'job_number': job_number, 'job_name': job.job_name})
                            if job.is_labor_done is False:
                                # job.is_labor_done = True
                                # job.labor_done_Date = date.today()
                                # job.save()
                                needs_to_be_labor_done.append({'job_number': job_number, 'job_name': job.job_name})
                        else:
                            if job.is_labor_done is True:
                                needs_to_not_be_labor_done.append({'job_number': job_number, 'job_name': job.job_name})
                    else:
                        not_found.append(job_number)
                a = a + 1
                if a > 200000:
                    break
            for x in Jobs.objects.filter(is_closed=False): #scroll through open jobs in Trinity
                if x.superintendent:
                    trinitysuper = x.superintendent.first_name
                    mcsuper = "couldn't find"
                    for y in superintendents:
                        if y['job_number'] == x.job_number:
                            mcsuper = y['super']
                            if mcsuper == 'Sub' and not trinitysuper == 'Victor':
                                superintendenterrors.append('Job ' + str(
                                    x.job_number) + ' has ' + mcsuper + ' in MC, but ' + trinitysuper + ' in Trinity')
                            elif mcsuper == 'Steve' and not trinitysuper == 'Steve':
                                superintendenterrors.append('Job ' + str(
                                    x.job_number) + ' has ' + mcsuper + ' in MC, but ' + trinitysuper + ' in Trinity')
                            elif mcsuper == 'Ed' and not trinitysuper == 'Edward':
                                superintendenterrors.append('Job ' + str(
                                    x.job_number) + ' has ' + mcsuper + ' in MC, but ' + trinitysuper + ' in Trinity')
                            break
                else:
                    for y in superintendents:
                        if y['job_number'] == x.job_number:
                            if y['super']:
                                superintendenterrors.append('Job ' + str(x.job_number) + ' has ' + y[
                                    'super'] + ' in MC, but no Super in Trinity')
                            break
                if x.job_number not in open_jobs:
                    if Inventory.objects.filter(job_number=x, is_closed=False).exists():
                        closed_but_equipment.append({'job_number': x.job_number, 'job_name': x.job_name})
                        message = "Job: " + x.job_name + " is closed. The following equipment is assigned to the job and must be returned immediately!\n "
                        recipients = ["admin1@gerloffpainting.com", "admin2@gerloffpainting.com",
                                      "warehouse@gerloffpainting.com", "joe@gerloffpainting.com"]
                        if x.superintendent.email is None:
                            message = message + "\n No email address for " + str(x.superintendent)
                        else:
                            recipients.append(x.superintendent.email)
                        for y in Inventory.objects.filter(job_number=x, is_closed=False):
                            if y.number:
                                message = message + "\n -" + y.item + " GP Number #" + y.number
                            else:
                                message = message + "\n -" + y.item + " -No GP Number! "
                        try:
                            Email.sendEmail("Closed Job - " + x.job_name, message,
                                            recipients, False)
                        except:
                            error_message += message
                    if Subcontracts.objects.filter(is_closed=False, job_number=x).exists():
                        closed_but_subs.append({'job_number': x.job_number, 'job_name': x.job_name})
                        message = "Job: " + x.job_name + " cannot be closed. The following subcontracts are still open!\n "
                        recipients = ["admin1@gerloffpainting.com", "admin2@gerloffpainting.com",
                                      "joe@gerloffpainting.com"]
                        for y in Subcontracts.objects.filter(is_closed=False, job_number=x):
                            if y.po_number:
                                message += "\n -" + y.subcontractor.company + " - PO# " + y.po_number + "! "
                            else:
                                message += "\n -" + y.subcontractor.company + " - PO# N/A!"
                        try:
                            Email.sendEmail("Closed Job Error- " + x.job_name, message,
                                            recipients, False)
                        except:
                            error_message += message

                    else:
                        needs_to_be_closed.append({'job_number': x.job_number, 'job_name': x.job_name})
                        x.is_closed = True
                        x.save()
    send_data['closed_but_subs']= closed_but_subs
    send_data['closed_but_equipment'] =closed_but_equipment
    send_data['need_to_be_labor_done'] =needs_to_be_labor_done
    send_data['needs_to_not_be_labor_done'] = needs_to_not_be_labor_done
    send_data['needs_to_be_opened'] = needs_to_be_opened
    send_data['needs_to_be_closed'] = needs_to_be_closed
    send_data['not_found'] = not_found
    send_data['employees'] = Employees.objects.filter(user__isnull=True, active=True)
    send_data['subs'] = Subcontractors.objects.filter(is_inactive=False)
    send_data['error_message'] = "These Emails Failed to send. Please let everyone know. " + error_message
    send_data['superintendenterrors'] = superintendenterrors
    return render(request, 'multi_use_page.html', send_data)


@login_required(login_url='/accounts/login')
def upload_new_job(request):
    if request.method == 'POST':
        if 'upload_file' in request.FILES:
            fileitem = request.FILES['upload_file']
            fn = os.path.basename(fileitem.name)
            fn2 = os.path.join(settings.MEDIA_ROOT, "job_upload", "Temp.xlsx")
            open(fn2, 'wb').write(fileitem.file.read())
            send_data = {}

            wb_obj = openpyxl.load_workbook(filename=request.FILES['upload_file'].file)
            sheet_obj = wb_obj["Data"]
            client_name = sheet_obj.cell(row=16, column=2).value
            pm_name = sheet_obj.cell(row=19, column=2).value
            send_data['job_name'] = sheet_obj.cell(row=38, column=2).value
            send_data['job_number'] = sheet_obj.cell(row=25, column=2).value
            send_data['new_client'] = client_name
            send_data['new_pm_name'] = sheet_obj.cell(row=19, column=2).value
            send_data['new_pm_phone'] = sheet_obj.cell(row=20, column=2).value
            send_data['new_pm_email'] = sheet_obj.cell(row=21, column=2).value
            send_data['new_super_name'] = sheet_obj.cell(row=22, column=2).value
            send_data['new_super_phone'] = sheet_obj.cell(row=23, column=2).value
            send_data['new_super_email'] = sheet_obj.cell(row=24, column=2).value
            if Clients.objects.filter(company=client_name).exists():
                send_data['client'] = Clients.objects.get(company=client_name)
                send_data['pms_filter'] = ClientEmployees.objects.filter(id=Clients.objects.get(company=client_name))
                if ClientEmployees.objects.filter(name=pm_name, id__company=client_name).exists():
                    send_data['pm'] = ClientEmployees.objects.get(name=pm_name, id__company=client_name)
                super_name = sheet_obj.cell(row=22, column=2).value
                if ClientEmployees.objects.filter(name=super_name, id__company=client_name).exists():
                    send_data['super'] = ClientEmployees.objects.get(name=super_name, id__company=client_name)
            if Employees.objects.filter(first_name=sheet_obj.cell(row=39, column=2).value).exists():
                send_data['estimator'] = Employees.objects.get(first_name=sheet_obj.cell(row=39, column=2).value)
            else:
                send_data['estimator_name'] = sheet_obj.cell(row=39, column=2).value
            send_data['all_clients'] = Clients.objects.all()
            send_data['all_employees'] = Employees.objects.exclude(job_title__description="Painter")
            send_data['data'] = json.dumps(list(ClientEmployees.objects.values('name', 'id', 'person_pk')),
                                           cls=DjangoJSONEncoder)
            return render(request, 'job_upload.html', send_data)
        if 'book_job' in request.POST:
            wb_obj = openpyxl.load_workbook(os.path.join(settings.MEDIA_ROOT, "job_upload", "Temp.xlsx"))
            sheet_obj = wb_obj["Data"]

            if request.POST['select_company'] == 'add_new':
                client = Clients.objects.create(company=request.POST['new_client'],
                                                bid_email=request.POST['new_client_bid_email'],
                                                phone=request.POST['new_client_phone'])

            else:
                client = Clients.objects.get(id=request.POST['select_company'])
                client.company = request.POST['new_client']
                client.bid_email = request.POST['new_client_bid_email']
                client.phone = request.POST['new_client_phone']
                client.save()
            if request.POST['select_pm'] == 'add_new':
                client_pm = ClientEmployees.objects.create(id=client, name=request.POST['new_pm'],
                                                           phone=request.POST['new_pm_phone'],
                                                           email=request.POST['new_pm_email'])

            else:
                client_pm = ClientEmployees.objects.get(person_pk=request.POST['select_pm'])
                client_pm.name = request.POST['new_pm']
                client_pm.phone = request.POST['new_pm_phone']
                client_pm.email = request.POST['new_pm_email']
                client_pm.save()

            gp_estimator = Employees.objects.get(id=request.POST['select_gpestimator'])
            job_number = sheet_obj.cell(row=25, column=2).value
            job_name = sheet_obj.cell(row=38, column=2).value
            address = sheet_obj.cell(row=40, column=2).value
            city = sheet_obj.cell(row=41, column=2).value
            state = sheet_obj.cell(row=256, column=2).value
            if sheet_obj.cell(row=42, column=2).value == "Yes":
                is_on_base = True
            else:
                is_on_base = False
            if sheet_obj.cell(row=47, column=2).value == "Yes":
                is_t_m_job = True
            else:
                is_t_m_job = False
            if sheet_obj.cell(row=3, column=2).value == "Yes":
                contract_status = 1  # received
            elif sheet_obj.cell(row=1, column=2).value == "Yes":
                contract_status = 2  # not received
            else:
                contract_status = 3  # not needed
            if sheet_obj.cell(row=6, column=2).value == "Yes":
                insurance_status = 1
            elif sheet_obj.cell(row=4, column=2).value == "Yes":
                insurance_status = 2
            else:
                insurance_status = 3

            start_date = sheet_obj.cell(row=33, column=2).value
            job = Jobs.objects.create(job_number=job_number, job_name=job_name, address=address, city=city,
                                      state=state,
                                      is_on_base=is_on_base, is_t_m_job=is_t_m_job, contract_status=contract_status,
                                      insurance_status=insurance_status, client=client, start_date=start_date,
                                      status="Open", booked_date=date.today(), client_Pm=client_pm,
                                      booked_by=request.user.first_name + " " + request.user.last_name,
                                      estimator=gp_estimator)

            if request.POST['select_super'] != "not_sure":
                if 'duplicate' in request.POST:
                    job.client_Super = client_pm
                elif request.POST['select_super'] == 'add_new':
                    job.client_Super = ClientEmployees.objects.create(id=client, name=request.POST['new_super'],
                                                                      phone=request.POST['new_super_phone'],
                                                                      email=request.POST['new_super_email'])
                elif request.POST['select_super'] == 'duplicate':
                    job.client_Super = client_pm
                else:
                    job.client_Super = ClientEmployees.objects.get(person_pk=request.POST['select_super'])

            if sheet_obj.cell(row=44, column=2).value: job.brush_role = sheet_obj.cell(row=44, column=2).value
            if sheet_obj.cell(row=45, column=2).value: job.spray_scale = sheet_obj.cell(row=45, column=2).value
            if sheet_obj.cell(row=48, column=2).value: job.t_m_nte_amount = sheet_obj.cell(row=48, column=2).value
            if sheet_obj.cell(row=74, column=2).value: job.contract_amount = sheet_obj.cell(row=74, column=2).value
            if sheet_obj.cell(row=76, column=2).value: job.painting_budget = sheet_obj.cell(row=76, column=2).value
            if sheet_obj.cell(row=77, column=2).value: job.wallcovering_budget = sheet_obj.cell(row=77, column=2).value
            if sheet_obj.cell(row=43, column=2).value: job.is_wage_scale = True
            if sheet_obj.cell(row=57, column=2).value: job.special_paint_needed = True
            if sheet_obj.cell(row=29, column=2).value: job.has_paint = True
            if sheet_obj.cell(row=30, column=2).value: job.has_wallcovering = True
            if sheet_obj.cell(row=52, column=2).value:
                job.submittals_needed = False
            else:
                job.submittals_needed = True
            if sheet_obj.cell(row=18, column=2).value: job.po_number = sheet_obj.cell(row=18, column=2).value
            if sheet_obj.cell(row=36, column=2).value: job.is_off_hours = True
            job.start_date_checked = date.today()
            job.save()
            temp_note = "New Job Booked By: " + request.user.first_name + " " + request.user.last_name
            if sheet_obj.cell(row=37, column=2).value: temp_note = temp_note + ": " + sheet_obj.cell(row=37,
                                                                                                     column=2).value
            if sheet_obj.cell(row=60, column=2).value: temp_note = temp_note + ": " + sheet_obj.cell(row=60,
                                                                                                     column=2).value

            JobNotes.objects.create(job_number=job,
                                    note="Start Date at Booking: " + start_date.strftime(
                                        "%m/%d/%Y") + " " + sheet_obj.cell(row=258,
                                                                           column=2).value,
                                    type="auto_start_date_note", date=date.today(),
                                    user=Employees.objects.get(user=request.user))
            JobNotes.objects.create(job_number=job,
                                    note=temp_note,
                                    type="auto_booking_note", date=date.today(),
                                    user=Employees.objects.get(user=request.user))
            email_body = "New Job Booked \n" + job.job_number + "\n" + job.job_name + "\n" + job.client.company
            try:
                Email.sendEmail("New Job - " + job.job_name, email_body,
                                ['admin1@gerloffpainting.com', 'admin2@gerloffpainting.com', 'joe@gerloffpainting.com'],
                                False)
                success = True
            except:
                success = False


            job_state = "Virginia"
            if job.state == "NC":
                job_state="North Carolina"
            payload = {
                #"job_name": job.job_name,
                "job_name": job.job_name,
                "job_number": job.job_number,
                "address": job.address,
                "city": job.city,
                "state": job_state,
                "country": "United States",

            }

            try:
                requests.post(
                    settings.ZAPIER_CREATE_JOB_WEBHOOK,
                    json=payload,
                    timeout=5,
                )
            except requests.RequestException:
                print("PUMPKIN ERROR")
                pass

            return render(request, "upload_new_job.html")
    return render(request, "upload_new_job.html")


@login_required(login_url='/accounts/login')
def jobs_home(request):
    send_data = {}
    if request.method == 'GET':
        if 'search' in request.GET: send_data['search_exists'] = request.GET['search']  # jobname
        if 'search2' in request.GET:
            send_data['search2_exists'] = request.GET['search2']  # super name
            if request.GET['search2'] != 'ALL' and request.GET['search2'] != 'UNASSIGNED':
                send_data['selected_supername'] = Employees.objects.get(
                    id=request.GET['search2']).first_name + " " + Employees.objects.get(
                    id=request.GET['search2']).last_name
        if 'search3' in request.GET: send_data['search3_exists'] = request.GET['search3']  # open only
        if 'search4' in request.GET: send_data['search4_exists'] = request.GET['search4']  # gc name
        if 'search5' in request.GET: send_data['search5_exists'] = request.GET['search5']  # upcoming only
        if 'search6' in request.GET: send_data['search6_exists'] = request.GET['search6']  # unassigned
        if 'search7' in request.GET: send_data['search7_exists'] = request.GET['search7']  # labor done
    search_jobs = JobsFilter(request.GET, queryset=Jobs.objects.filter())
    send_data['search_jobs'] = JobsFilter(request.GET, queryset=Jobs.objects.filter())
    send_data['jobstable'] = search_jobs.qs.order_by('start_date')
    # RequestConfig(request).configure(jobstable)
    # RequestConfig(request, paginate=False).configure(jobstable)
    send_data['has_filter'] = any(field in request.GET for field in set(search_jobs.get_fields()))
    send_data['supers'] = Employees.objects.filter(job_title__description="Superintendent", active=True)
    send_data['tickets'] = ChangeOrders.objects.filter(job_number__is_closed=False, is_t_and_m=True,
                                                       is_ticket_signed=False)
    send_data['open_cos'] = ChangeOrders.objects.filter(job_number__is_closed=False, is_closed=False,
                                                        is_approved=False) & ChangeOrders.objects.filter(
        is_t_and_m=False) | ChangeOrders.objects.filter(is_t_and_m=True, is_ticket_signed=True)
    send_data['approved_cos'] = ChangeOrders.objects.filter(job_number__is_closed=False, is_closed=False,
                                                            is_approved=True)
    send_data['equipment'] = Inventory.objects.filter(job_number__is_closed=False, is_closed=False).order_by(
        'inventory_type')
    send_data['rentals'] = Rentals.objects.filter(job_number__is_closed=False, off_rent_number__isnull=True)
    wallcovering2 = Wallcovering.objects.filter(job_number__is_closed=False)
    wc_not_ordereds = []
    for x in wallcovering2:
        if x.orderitems1.count() > 0:
            print(x)
        else:
            wc_not_ordereds.append(x)
    send_data['wc_not_ordereds'] = wc_not_ordereds
    send_data['wc_ordereds'] = OrderItems.objects.filter(wallcovering__job_number__is_closed=False,
                                                         is_satisfied=False)
    send_data['packages'] = Packages.objects.filter(delivery__order__job_number__is_closed=False)
    send_data['deliveries'] = OutgoingItem.objects.filter(outgoing_event__job_number__is_closed=False)
    send_data['submittals'] = Submittals.objects.filter(job_number__is_closed=False)
    subcontracts = []
    for x in Subcontracts.objects.filter(job_number__is_closed=False, is_closed=False):
        total_contract = "{:,}".format(int(x.total_contract_amount()))
        percent_complete = format(x.percent_complete(), ".0%")
        subcontracts.append({'id': x.id, 'po_number': x.po_number, 'subcontractor': x.subcontractor.company,
                             'total_contract': total_contract, 'percent_complete': percent_complete})
    send_data['subcontracts'] = subcontracts
    send_data['jobs'] = 'ALL'
    return render(request, "jobs_home.html", send_data)

def activate_sub_job(request,jobnumber):
    selected_job=Jobs.objects.get(job_number=jobnumber)
    if selected_job.is_painting_subbed:
        selected_job.is_painting_subbed = False
    else:
        selected_job.is_painting_subbed = True
    selected_job.save()
    return redirect('job_page', jobnumber=jobnumber)

@login_required(login_url='/accounts/login')
def job_page(request, jobnumber):
    go_to_pickup = False
    selectedjob = Jobs.objects.get(job_number=jobnumber)
    send_data = {}
    if Email_Errors.objects.filter(user=request.user.first_name + " " + request.user.last_name).exists():
        send_data['error_message']= Email_Errors.objects.filter(user=request.user.first_name + " " + request.user.last_name).last().error
    Email_Errors.objects.filter(user=request.user.first_name + " " + request.user.last_name).delete()
    if PickupRequest.objects.filter(job_number=selectedjob, is_closed=False, confirmed=True).exists():
        pickup_request = PickupRequest.objects.get(job_number=selectedjob, is_closed=False, confirmed=True)
        send_data['pickup_exists'] = True
        if pickup_request.remove_trash:
            send_data['trash_pickup_requested'] = True
    if request.method == 'POST':
        if 'add_competent_person' in request.POST:
            selected_competent_person_id = request.POST['add_competent_person']
            selected_competent_person = Employees.objects.get(id=selected_competent_person_id)
            if not Competent_Persons.objects.filter(employee=selected_competent_person,job=selectedjob).exists():
                Competent_Persons.objects.create(employee=selected_competent_person,job=selectedjob,date=date.today())
                message = "Competent Person " + selected_competent_person.first_name + " " + selected_competent_person.last_name + " has been assigned to " + selectedjob.job_name
                recipients = ["skip@gerloffpainting.com","bridgette@gerloffpainting.com"]
                if selectedjob.superintendent:
                    if selectedjob.superintendent.email:
                        recipients.append(selectedjob.superintendent.email)
                    else:
                        recipients.append("victor@gerloffpainting.com")
                else:
                    recipients.append("victor@gerloffpainting.com")
                try:
                    Email.sendEmail("Competent Person Assigned", message, recipients, False)
                    send_data['error_message'] = "Email announcing competent person was successfully sent. "
                except:
                    send_data['error_message'] = "ERROR! Your Email announcing competent person was NOT sent. Please tell the super. "
        if 'safety_packet' in request.POST:
            if 'safety_packet_sent' in request.POST:
                selectedjob.has_safety_packet_been_sent = True
            else:
                selectedjob.has_safety_packet_been_sent = False
            selectedjob.save()
        if 'undo_redo' in request.POST:
            selectedEmployeeIds = request.POST.getlist('undo_redo')
            EmployeeJob.objects.filter(job=jobnumber).delete()
            for id in selectedEmployeeIds:
                employee = Employees.objects.get(id=id)
                newSelectedEmployee = EmployeeJob.objects.create(employee=employee, job=selectedjob)
                newSelectedEmployee.save()
        if 'start_date' in request.POST:
            return redirect('job_page', jobnumber=jobnumber)
        if 'select_rental' in request.POST:
            selected_rental = Rentals.objects.get(id=request.POST['select_rental'])
            selected_rental.requested_off_rent = True
            selected_rental.save()
            RentalNotes.objects.create(rental=selected_rental, date=date.today(),
                                       user=Employees.objects.get(user=request.user),
                                       note="Please call off-rent. " + request.POST['off_rent_note'])
            message = "Please call off this rental. " + selected_rental.item + ". From Job -" + selected_rental.job_number.job_name + "\n " + \
                      request.POST['off_rent_note']
            try:
                Email.sendEmail("Call Off Rent", message, ["warehouse@gerloffpainting.com"], False)
                send_data['error_message'] = "Email requesting off rent was successfully sent. "
            except:
                send_data['error_message'] = "ERROR! Your Email requesting off rent was NOT sent. Please call the warehouse. "
        if 'select_status' in request.POST:
            if request.POST['select_status'] == 'nothing_done':
                message = "Labor is not done."
                if selectedjob.is_labor_done == True:
                    try:
                        Email.sendEmail("Labor not done - " + selectedjob.job_name,
                                        "Per " + request.user.first_name + " " + request.user.last_name + "- Labor is not done. Please make sure to Un-Click the LABOR DONE box in management console. " +
                                        request.POST['closed_note'],
                                        ['bridgette@gerloffpainting.com',
                                         'victor@gerloffpainting.com'], False)
                        send_data['error_message'] = "Email about labor not done was successfully sent. "
                    except:
                        send_data['error_message'] = "ERROR! Email about labor not done was NOT sent. Please call the office."
                selectedjob.labor_done_Date = None
                selectedjob.is_waiting_for_punchlist = False
                selectedjob.is_labor_done = False
            if request.POST['select_status'] == 'waiting_for_punchlist':
                message = "Waiting for punchlist."
                selectedjob.labor_done_Date = None
                selectedjob.is_waiting_for_punchlist = True
                selectedjob.is_labor_done = False
                selectedjob.save()
                if Inventory.objects.filter(job_number=selectedjob, is_closed=False):
                    if PickupRequest.objects.filter(job_number=selectedjob, is_closed=False,
                                                    all_items=True).exists():
                        go_to_pickup = False
                    else:
                        go_to_pickup = True
            if request.POST['select_status'] == 'done_done':
                message = "Labor is 100% done."
                try:
                    Email.sendEmail("Labor Done - " + selectedjob.job_name,
                                    "Per " + request.user.first_name + " " + request.user.last_name + "- Labor is 100% Done. Please make sure to Click the Labor Done button in Management Console. " +
                                    request.POST['closed_note'],
                                    ['admin2@gerloffpainting.com',
                                     'bridgette@gerloffpainting.com',
                                     'victor@gerloffpainting.com'], False)
                    send_data['error_message'] = "Labor Complete Email was successfully sent. "
                except:
                    send_data['error_message'] = "ERROR! Your email about job being complete was NOT sent.  Please call the office. "
                selectedjob.labor_done_Date = date.today()
                selectedjob.is_waiting_for_punchlist = True
                selectedjob.is_labor_done = True
                selectedjob.save()
                if Inventory.objects.filter(job_number=selectedjob, is_closed=False):
                    if PickupRequest.objects.filter(job_number=selectedjob, is_closed=False,
                                                    all_items=True).exists():
                        go_to_pickup = False
                    else:
                        go_to_pickup = True
            selectedjob.save()
            JobNotes.objects.create(job_number=selectedjob,
                                    note=message + " " + request.POST['closed_note'], type="employee_note",
                                    user=Employees.objects.get(user=request.user), date=date.today())

        if 'add_note' in request.POST:
            JobNotes.objects.create(job_number=selectedjob,
                                    note=request.POST['add_note'], type="employee_note",
                                    user=Employees.objects.get(user=request.user), date=date.today())
            send_data['notes_open_button'] = True
        if 'submit_pm' in request.POST:
            if request.POST['select_pm'] == 'add_new':
                selectedjob.client_Pm = ClientEmployees.objects.create(id=selectedjob.client,
                                                                       name=request.POST['new_pm'],
                                                                       phone=request.POST['new_pm_phone'],
                                                                       email=request.POST['new_pm_email'])
                selectedjob.save()
            else:
                selectedjob.client_Pm = ClientEmployees.objects.get(person_pk=request.POST['select_pm'])
                selectedjob.save()
                selectedjob.client_Pm.email = request.POST['pm_email']
                selectedjob.client_Pm.phone = request.POST['pm_phone']
                selectedjob.client_Pm.name = request.POST['pm_name']
                selectedjob.client_Pm.save()
        if 'submit_super' in request.POST:
            if request.POST['select_super'] == 'add_new':
                selectedjob.client_Super = ClientEmployees.objects.create(id=selectedjob.client,
                                                                          name=request.POST['new_super'],
                                                                          phone=request.POST['new_super_phone'],
                                                                          email=request.POST['new_super_email'])
                selectedjob.save()
            elif request.POST['select_super'] != 'not_sure':
                selectedjob.client_Super = ClientEmployees.objects.get(person_pk=request.POST['select_super'])
                selectedjob.save()
                selectedjob.client_Super.email = request.POST['super_email']
                selectedjob.client_Super.phone = request.POST['super_phone']
                selectedjob.client_Super.save()
            elif selectedjob.client_Super:
                selectedjob.client_Super = None
                selectedjob.save()
        if 'submit_client' in request.POST:
            selectedjob.client.bid_email = request.POST['client_bid_email']
            selectedjob.client.phone = request.POST['client_phone']
            selectedjob.save()
    send_data['client_employees'] = ClientEmployees.objects.filter(id=selectedjob.client)
    send_data['job'] = selectedjob
    if selectedjob.labor_done_Date:
        short_year = selectedjob.labor_done_Date.strftime("%y")
        short_mth = selectedjob.labor_done_Date.strftime("%m")
        short_day = selectedjob.labor_done_Date.strftime("%d")
        send_data['labor_done_date'] = short_mth + "-" + short_day + "-" + short_year
    if selectedjob.current_contract_amount() != 0:
        contract_amount = int(selectedjob.current_contract_amount())
        contract_amount = ('{:,}'.format(contract_amount))
    else:
        contract_amount = "T&M"
    send_data['contract_amount'] = contract_amount
    if selectedjob.contract_amount:
        send_data['orig_contract_amount'] = ('{:,}'.format(selectedjob.contract_amount))
    else:
        send_data['orig_contract_amount'] = None
    send_data['pending_count'] = selectedjob.count_pending_changes()
    send_data['count_approved_changes'] = selectedjob.count_approved_changes()
    send_data['pending_co_amount'] = ('{:,}'.format(selectedjob.pending_co_amount()))
    send_data['approved_co_amount'] = ('{:,}'.format(selectedjob.approved_co_amount()))
    tickets_not_done = ChangeOrders.objects.filter(job_number=selectedjob, is_t_and_m=True,
                                                   is_ticket_signed=False,
                                                   is_old_form_printed=False, ewt__isnull=True).order_by('cop_number')
    send_data['tickets_not_done'] = tickets_not_done
    send_data['tickets_not_done_count'] = tickets_not_done.count()
    # tickets_not_signed = ChangeOrders.objects.filter(job_number=selectedjob, is_t_and_m=True,
    #                                                               is_ticket_signed=False,
    #                                                               is_old_form_printed=True)
    tickets_not_signed = ChangeOrders.objects.filter(job_number=selectedjob, is_t_and_m=True,
                                                     is_ticket_signed=False,
                                                     is_old_form_printed=True) | ChangeOrders.objects.filter(
        job_number=selectedjob, is_t_and_m=True, is_ticket_signed=False, ewt__isnull=False).order_by('cop_number')
    send_data['tickets_not_signed'] = tickets_not_signed
    send_data['tickets_not_signed_count'] = tickets_not_signed.count()
    tickets_not_sent = ChangeOrders.objects.filter(job_number=selectedjob, is_t_and_m=True,
                                                   is_ticket_signed=True, date_sent__isnull=True).order_by('cop_number')
    send_data['tickets_not_sent'] = tickets_not_sent
    send_data['tickets_not_sent_count'] = tickets_not_sent.count()
    open_cos = ChangeOrders.objects.filter(job_number=selectedjob, is_closed=False,
                                           is_approved=False, date_sent__isnull=False).order_by('cop_number')
    send_data['open_cos'] = open_cos
    send_data['open_cos_count'] = open_cos.count()
    informal_cos = ChangeOrders.objects.filter(job_number=selectedjob, is_closed=False,
                                               is_approved_to_bill=False, is_approved=True).order_by('cop_number')
    send_data['informal_cos'] = informal_cos
    send_data['informal_cos_count'] = informal_cos.count()
    approved_cos = ChangeOrders.objects.filter(job_number=selectedjob, is_closed=False,
                                               is_approved_to_bill=True).order_by('cop_number')
    send_data['approved_cos'] = approved_cos
    send_data['approved_cos_count'] = approved_cos.count()
    send_data['equipments'] = Inventory.objects.filter(job_number=selectedjob, is_closed=False).order_by(
        'inventory_type')
    send_data['rentals'] = Rentals.objects.filter(job_number=selectedjob, off_rent_number__isnull=True, is_closed=False)
    send_data['formals'] = selectedjob.formals()
    if Inventory.objects.filter(job_number=selectedjob, is_closed=False).order_by('inventory_type').exists():
        send_data['has_equipment'] = True
    if Rentals.objects.filter(job_number=selectedjob, off_rent_number__isnull=True).exists():
        send_data['has_rentals'] = True
    send_data['wallcovering2'] = Wallcovering.objects.filter(job_number=selectedjob)
    send_data['wc_not_ordereds'] = Wallcovering.objects.filter(job_number=selectedjob,
                                                               orderitems1__isnull=True)
    send_data['wc_ordereds'] = OrderItems.objects.filter(order__job_number=selectedjob, is_satisfied=False)
    send_data['packages'] = Packages.objects.filter(delivery__order__job_number=selectedjob)
    send_data['deliveries'] = OutgoingItem.objects.filter(outgoing_event__job_number=selectedjob)
    send_data['submittals'] = Submittals.objects.filter(job_number=selectedjob)
    subcontracts = []
    for x in Subcontracts.objects.filter(job_number=selectedjob):
        total_contract = "{:,}".format(int(x.total_contract_amount()))
        percent_complete = format(x.percent_complete(), ".0%")
        subcontracts.append({'id': x.id, 'po_number': x.po_number, 'subcontractor': x.subcontractor.company,
                             'total_contract': total_contract, 'percent_complete': percent_complete})
    send_data['subcontracts'] = subcontracts
    all_notes = JobNotesFilter(request.GET, queryset=JobNotes.objects.filter(job_number=selectedjob))
    send_data['all_notes'] = all_notes
    send_data['filtered_notes'] = all_notes.qs
    send_data['has_filter'] = any(field in request.GET for field in set(all_notes.get_fields()))
    if request.method == 'GET':
        made_already = False
        if 'admin' in request.GET:
            send_data['notes_open_button'] = True
            notes = JobNotes.objects.filter(job_number=selectedjob,
                                            type="auto_booking_note") | JobNotes.objects.filter(
                job_number=selectedjob, type="auto_misc_note")
            made_already = True
            send_data['admin'] = 'admin'
            send_data['notes_open_button'] = True
        if 'start' in request.GET:
            send_data['notes_open_button'] = True
            send_data['start'] = 'start'
            if made_already == False:
                notes = JobNotes.objects.filter(job_number=selectedjob, type="auto_start_date_note")
                made_already = True
            else:
                notes = notes | JobNotes.objects.filter(job_number=selectedjob, type="auto_start_date_note")
                made_already = True
        if 'field' in request.GET:
            send_data['notes_open_button'] = True
            send_data['field'] = 'field'
            if made_already == False:
                notes = JobNotes.objects.filter(job_number=selectedjob,
                                                type="employee_note") | JobNotes.objects.filter(
                    job_number=selectedjob, type="daily_report")
                made_already = True
            else:
                notes = notes | JobNotes.objects.filter(job_number=selectedjob,
                                                        type="employee_note") | JobNotes.objects.filter(
                    job_number=selectedjob, type="daily_report")
                made_already = True
        if 'change_order' in request.GET:
            send_data['notes_open_button'] = True
            send_data['change_order'] = 'change_order'
            if made_already == False:
                notes = JobNotes.objects.filter(job_number=selectedjob, type="auto_co_note")
                made_already = True
            else:
                notes = notes | JobNotes.objects.filter(job_number=selectedjob, type="auto_co_note")
                made_already = True
        if 'submittal' in request.GET:
            send_data['notes_open_button'] = True
            send_data['submittal'] = 'submittal'
            if made_already == False:
                notes = JobNotes.objects.filter(job_number=selectedjob, type="auto_submittal_note")
                made_already = True
            else:
                notes = notes | JobNotes.objects.filter(job_number=selectedjob, type="auto_submittal_note")
                made_already = True
        if made_already == False:
            notes = JobNotes.objects.filter(job_number=selectedjob)
    else:
        notes = JobNotes.objects.filter(job_number=selectedjob)
    send_data['notes'] = notes.order_by('date')
    send_data['supers'] = Employees.objects.filter(job_title__description="Superintendent", active=True)
    if go_to_pickup:
        return redirect('request_pickup', jobnumber=selectedjob.job_number, item='ALL', pickup='ALL', status='ALL')
    else:
        employees = Employees.objects.filter(job_title=1)
        filteredEmployees = []
        selectedEmployees = EmployeeJob.objects.filter(job=selectedjob.job_number)
        for employee in employees:
            found = False
            for selectedEmployee in selectedEmployees:
                if selectedEmployee.employee.id == employee.id:
                    found = True
            if found == False:
                filteredEmployees.append(employee)
        send_data['employees'] = filteredEmployees
        send_data['selectedEmployees'] = selectedEmployees
        send_data['painters'] = Employees.objects.filter(job_title__description="Painter",active=True)
        send_data['competent_persons'] = Competent_Persons.objects.filter(job=selectedjob)
        return render(request, 'job_page.html', send_data)


@login_required(login_url='/accounts/login')
def book_new_job(request):
    allclients = Clients.objects.order_by('company')
    estimators = Employees.objects.exclude(job_title__description='Painter')
    superintendents = Employees.objects.filter(job_title__description='Superintendent')
    pms = ClientEmployees.objects.values('name', 'id', 'person_pk')
    prices_json = json.dumps(list(pms), cls=DjangoJSONEncoder)

    return render(request, 'book_new_job.html',
                  {'data': prices_json, 'allclients': allclients, 'superintendents': superintendents,
                   'estimators': estimators})


@login_required(login_url='/accounts/login')
def register(request):
    if request.method == 'POST':
        checklist = []
        if request.POST['job_number'] == "":
            if JobNumbers.objects.latest("id").number == "9990":
                new_letter = chr(ord(JobNumbers.objects.latest("id").letter) + 1)
                new_job_number = JobNumbers.objects.create(letter=new_letter, number="0010")
                job_number = new_job_number.letter + new_job_number.number
            else:
                new_number = int(JobNumbers.objects.latest("id").number) + 10
                new_number_convert = str(new_number)
                if new_number < 100:
                    new_job_number = JobNumbers.objects.create(letter=JobNumbers.objects.latest("id").letter,
                                                               number="00" + new_number_convert)
                elif new_number < 1000:
                    new_job_number = JobNumbers.objects.create(letter=JobNumbers.objects.latest("id").letter,
                                                               number="0" + new_number_convert)
                else:
                    new_job_number = JobNumbers.objects.create(letter=JobNumbers.objects.latest("id").letter,
                                                               number=new_number_convert)
                new_job_number.save()
                job_number = new_job_number.letter + new_job_number.number
        else:
            job_number = request.POST['job_number']
        if request.POST['select_company'] == 'add_new':
            client = Clients.objects.create(company=request.POST['new_client'],
                                            bid_email=request.POST['new_client_bid_email'],
                                            phone=request.POST['new_client_phone'])

        else:
            client = Clients.objects.get(id=request.POST['select_company'])

        if request.POST['select_pm'] == 'use_below':
            client_pm = ClientEmployees.objects.create(id=client, name=request.POST['new_pm'],
                                                       phone=request.POST['new_pm_phone'],
                                                       email=request.POST['new_pm_email'])

        else:
            client_pm = ClientEmployees.objects.get(person_pk=request.POST['select_pm'])

        job = Jobs.objects.create(job_number=job_number, job_name=request.POST['job_name'],
                                  address=request.POST['address'], city=request.POST['city'],
                                  state=request.POST['state'],
                                  contract_status=request.POST['contract_status'],
                                  insurance_status=request.POST['insurance_status'], client=client, client_Pm=client_pm,
                                  start_date=request.POST['start_date'],
                                  status="Open", booked_date=date.today(),
                                  booked_by=request.user.first_name + " " + request.user.last_name,
                                  estimator=Employees.objects.get(id=request.POST['select_gpestimator']),
                                  notes=request.POST['email_job_note'], po_number=request.POST['po_number'])

        if request.POST['select_super'] == 'not_sure':
            checklist.append("get superintendent info")
        elif request.POST['select_super'] == 'use_below':
            job.client_Super = ClientEmployees.objects.create(id=client, name=request.POST['new_super'],
                                                              phone=request.POST['new_super_phone'],
                                                              email=request.POST['new_super_email'])

        else:
            job.client_Super = ClientEmployees.objects.get(person_pk=request.POST['select_super'])
        if request.POST['spray_scale'] != "": job.spray_scale = request.POST['spray_scale']
        if request.POST['brush_role'] != "": job.brush_role = request.POST['brush_role']
        if request.POST['t_m_nte_amount'] != "": job.t_m_nte_amount = request.POST['t_m_nte_amount']
        if request.POST['select_gpsuper'] != "not_sure": job.superintendent = Employees.objects.get(
            id=request.POST['select_gpsuper'])
        # add email message request.POST['email_job_note']
        if request.POST['contract_amount'] != "": contract_amount = request.POST['contract_amount']
        if request.POST['painting_budget'] != "": painting_budget = request.POST['painting_budget']
        if request.POST['wallcovering_budget'] != "": wallcovering_budget = request.POST['wallcovering_budget']
        if 'is_t_m_job' in request.POST: job.is_t_m_job = True
        if 'on_base2' in request.POST: job.is_on_base = True
        if 'is_wage_rate' in request.POST: job.is_wage_scale = True
        if 'is_bonded' in request.POST: job.is_bonded = True
        if 'has_special_paint' in request.POST: job.special_paint_needed = True
        if 'has_paint' in request.POST: job.has_paint = True
        if 'has_wallcovering' in request.POST: job.has_wallcovering = True
        if 'has_submittals' in request.POST:  job.submittals_needed = True
        job.start_date_checked = date.today()
        job.save()
        JobNotes.objects.create(job_number=job,
                                note="Start Date at Booking: " + job.start_date + " " + request.POST['date_note'],
                                type="auto_start_date_note", date=date.today(),
                                user=Employees.objects.get(user=request.user))
        JobNotes.objects.create(job_number=job,
                                note="New Job Booked By: " + request.user.first_name + " " + request.user.last_name + ": " +
                                     request.POST['email_job_note'],
                                type="auto_booking_note", date=date.today(),
                                user=Employees.objects.get(user=request.user))
        email_body = "New Job Booked \n" + job.job_number + "\n" + job.job_name + "\n" + job.client.company
        try:
            Email.sendEmail("New Job - " + job.job_name, email_body, ['joe@gerloffpainting.com'], False)
            success = True
        except:
            success = False
        job.save()
        # for x in checklist:
        #     checklist = Checklist(job_number=job_number, checklist_item=x, category="PM")
        #     checklist.save();
        response = redirect('/')
        return response


@login_required(login_url='/accounts/login')
class Checklist(models.Model): #as of 1/20/2026 joe doesnt think this is used
    id = models.BigAutoField(primary_key=True)
    category = models.CharField(null=True, max_length=1000)
    checklist_item = models.CharField(null=True, max_length=2000)
    is_closed = models.BooleanField(default=False)
    job_number = models.ForeignKey(Jobs, on_delete=models.PROTECT)
    notes = models.CharField(null=True, max_length=2500)
    job_start_date_from_schedule = models.DateField(null=True, blank=True)
    cop = models.BooleanField(default=False)
    cop_amount = models.DecimalField(
        max_digits=6, decimal_places=2, blank=True, null=True)
    cop_sent_date = models.DateField(null=True, blank=True)
    cop_number = models.IntegerField(default=0)
    is_ewt = models.BooleanField(default=False)
    ewt_date = models.DateField(null=True, blank=True)
    is_submittal = models.BooleanField(default=False)
    submittal_number = models.IntegerField(default=0)
    submittal_description = models.CharField(null=True, max_length=2000)
    submittal_date_sent = models.DateField(null=True, blank=True)
    wallcovering_order_date = models.DateField(null=True, blank=True)
    assigned = models.CharField(null=True, max_length=2000)

    def __str__(self):
        return f"{self.job_number} {self.checklist_item}"

def get_work_day(time_raw):
    """
    Determine the work day (date only) for this entry.
    Prefer start date; fall back to end date.
    """
    if time_raw:
        return parse_date(time_raw).date()
    if end_raw:
        return parse_date(end_raw).date()
    return None

@csrf_exempt
def clockshark_webhook(request):
    if request.headers.get("X-ZAPIER-SECRET") != settings.ZAPIER_SECRET:
        return JsonResponse({"error": "Unauthorized"}, status=401)

    if request.method != "POST":
        return JsonResponse({"error": "Invalid method"}, status=405)

    try:
        payload = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({"error": "Invalid JSON"}, status=400)
    success=False
    employee_first_name = payload.get("employee_first_name")
    employee_last_name = payload.get("employee_last_name")
    job_name = payload.get("job_name")
    start_raw = payload.get("start")   # may be missing on clock-out
    end_raw = payload.get("end")       # should exist on clock-out
    clock_in_time = parse_date(start_raw) if start_raw else None
    clock_out_time = parse_date(end_raw) if end_raw else None

    # Normalize to timezone-aware if needed (dateutil sometimes returns naive)
    if clock_in_time and timezone.is_naive(clock_in_time):
        #clock_in_time = timezone.make_aware(clock_in_time, timezone.get_current_timezone())
        clock_in_time = timezone.make_aware(
            clock_in_time,
            timezone.utc
        )
        print(clock_in_time)
    if clock_out_time and timezone.is_naive(clock_out_time):
        #clock_out_time = timezone.make_aware(clock_out_time, timezone.get_current_timezone())
        clock_out_time = timezone.make_aware(
            clock_out_time,
            timezone.utc
        )
    if clock_in_time:
        work_day = parse_date(start_raw).date()
    if clock_out_time:
        work_day = parse_date(end_raw).date()

    clockshark_id = f"{employee_first_name}|{employee_last_name}|{job_name}|{work_day}"
    job = Jobs.objects.filter(job_name=job_name).first()
    # ---- Case A: CLOCK IN event (has start, no end) ----
    if clock_in_time:
        if ClockSharkTimeEntry.objects.filter(clockshark_id=clockshark_id,clock_out__isnull=False,clock_in__isnull=True).exists():
            ClockSharkErrors.objects.create(clockshark_id=clockshark_id,job_name=job_name,employee_first_name=employee_first_name,employee_last_name=employee_last_name,work_day=work_day,clock_in=clock_in_time, error="already an entry with a clock out time, without a clock in time")
            entry = ClockSharkTimeEntry.objects.filter(clockshark_id=clockshark_id,clock_out__isnull=False,clock_in__isnull=True).first()
            ClockSharkErrors.objects.create(clockshark_id=entry.clockshark_id, job_name=entry.job_name,
                                            employee_first_name=entry.employee_first_name,
                                            employee_last_name=entry.employee_last_name, work_day=entry.work_day,
                                            clock_out=entry.clock_out,
                                            error="no clock in time")
            entry.delete()
            return JsonResponse({"status": "already an entry with clock out time, no clock in time"})
        elif ClockSharkTimeEntry.objects.filter(clockshark_id=clockshark_id,clock_out__isnull=True,clock_in__isnull=False).exists():
            ClockSharkErrors.objects.create(clockshark_id=clockshark_id,job_name=job_name,employee_first_name=employee_first_name,employee_last_name=employee_last_name,work_day=work_day,clock_in=clock_in_time, error="already an entry with a clock-in time, without a clock out time")
            entry = ClockSharkTimeEntry.objects.filter(clockshark_id=clockshark_id,clock_out__isnull=True,clock_in__isnull=False).first()
            ClockSharkErrors.objects.create(clockshark_id=clockshark_id, job_name=job_name,
                                            employee_first_name=employee_first_name,
                                            employee_last_name=employee_last_name, work_day=work_day,
                                            clock_in=entry.clock_in,
                                            error="another clock-in came in, before a clock out time")
            entry.delete()
            return JsonResponse({"status": "another clock-in came in, before a clock out time"})
        else:
            if not job:
                ClockSharkErrors.objects.create(clockshark_id=clockshark_id, job_name=job_name,
                                                employee_first_name=employee_first_name,
                                                employee_last_name=employee_last_name, work_day=work_day,
                                                clock_out=clock_out_time,
                                                error="can't find job")
                ClockSharkTimeEntry.objects.create(clockshark_id=clockshark_id, job_name=job_name,
                                               employee_first_name=employee_first_name,
                                               employee_last_name=employee_last_name, work_day=work_day,
                                               clock_in=clock_in_time)
                return JsonResponse({"status": "success. but couldn't find job"})

            #ClockSharkTimeEntry.objects.create(clockshark_id=clockshark_id,job_name=job_name,employee_first_name=employee_first_name,employee_last_name=employee_last_name,work_day=work_day,clock_in=clock_in_time,job=job)
            if job:
                ClockSharkTimeEntry.objects.create(clockshark_id=clockshark_id, job_name=job_name,
                                                   employee_first_name=employee_first_name,
                                                   employee_last_name=employee_last_name, work_day=work_day,
                                                   clock_in=clock_in_time, job=job)
                if not job.is_active:
                    if not JobNotes.objects.filter(job_number=job, note__contains="Changed Status to Active").exists():
                        job.is_active = True
                        job.save()
                        JobNotes.objects.create(job_number=job, note="Changed Status to Active From Clock Shark Import",
                                                type="auto_start_date_note", user=Employees.objects.filter().first(), date=date.today())
                return JsonResponse({"status": "clock_in_saved"})


    # ---- Case B: CLOCK OUT event  ----
    if clock_out_time:
        if ClockSharkTimeEntry.objects.filter(clockshark_id=clockshark_id,clock_out__isnull=False, clock_in__isnull=True).exists():
            ClockSharkErrors.objects.create(clockshark_id=clockshark_id,job_name=job_name, employee_first_name=employee_first_name,
                                            employee_last_name=employee_last_name, work_day=work_day,
                                            clock_out=clock_out_time,
                                            error="already an entry with a clock-out time, without a clock in time")
            entry = ClockSharkTimeEntry.objects.filter(clockshark_id=clockshark_id,clock_out__isnull=False, clock_in__isnull=True).first()
            ClockSharkErrors.objects.create(clockshark_id=clockshark_id,job_name=job_name, employee_first_name=employee_first_name,
                                            employee_last_name=employee_last_name, work_day=work_day,
                                            clock_out=entry.clock_out,
                                            error="no clock in time")
            entry.delete()
            return JsonResponse({"status": "no clock in before clock out"})
        else:
            if ClockSharkTimeEntry.objects.filter(clockshark_id=clockshark_id,clock_in__isnull=False,clock_out__isnull=True).exists():
                entry = ClockSharkTimeEntry.objects.filter(clockshark_id=clockshark_id, clock_in__isnull=False,clock_out__isnull=True).first()
                if clock_out_time > entry.clock_in:
                    entry.clock_out = clock_out_time
                    entry.save()
                    success=True
                else:
                    ClockSharkErrors.objects.create(job_name=job_name, employee_first_name=employee_first_name,
                                                    employee_last_name=employee_last_name, work_day=work_day,
                                                    clock_out=clock_out_time,
                                                    error="clock out time is less than clock in time")
                    return JsonResponse({"status": "clock out time is less than clock in time"})
            else:
                ClockSharkErrors.objects.create(job_name=job_name, employee_first_name=employee_first_name,
                                                employee_last_name=employee_last_name, work_day=work_day,
                                                clock_out=clock_out_time,
                                                error="couldn't find a clock in time")
                return JsonResponse({"status": "couldn't find a clock in time"})



        # Compute hours from the stored clock_in
        if success:
            delta = clock_out_time - entry.clock_in
            hours = Decimal(delta.total_seconds() / 3600).quantize(Decimal("0.01"))
            if hours >= Decimal("6.00"):
                hours -= Decimal("0.50")
                entry.lunch = Decimal("30.00")
            entry.hours = hours
            entry.save()
            return JsonResponse({"status": "clock_out_updated", "hours": str(hours)})

    return JsonResponse({"status": "ignored"})