from .models import *
from accounts.models import *
from changeorder.models import *
from console.misc import createfolder, Email
from console.models import *
from datetime import datetime,date,timedelta
from dateutil.parser import parse as parse_date
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from difflib import SequenceMatcher
from django.apps import apps
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import ForeignKey
from django.views.decorators.cache import never_cache
from django.contrib.auth.models import User, auth
from django.db import transaction
from django.db.models import Max
from django.http import HttpResponse, JsonResponse
from django.db.models import Q
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.utils import timezone
from employees.forms import SiriusUploadForm,ClockSharkUploadForm, ToolboxTalksUploadForm
from employees.models import *
from employees.views import (
    _apply_certification_display_descriptions,
    _create_standard_certification_custom_attributes,
    get_respirators_in_review,
    get_scheduled_toolbox_folder,
    get_uploaded_toolbox_file,
)
from jobs.models import *
from jobs.models import Jobs
from jobs.views import subtract_months
from openpyxl import load_workbook
from pathlib import Path
from rentals.models import *
from subcontractors.models import *
from subcontractors import toolbox_views as sub_toolbox
from submittals.models import *
from superintendent.models import *
from wallcovering.models import *
import changeorder.models
import csv
import datetime
import employees.models
import equipment.models
import jobs.models
import json
import openpyxl
import os
import os.path
import random
import re


CERTIFICATION_IMPORT_REQUIRED_HEADERS = {
    "employee": ["employee", "employee name", "name"],
    "subcontractor": ["subcontractor", "sub contractor", "company", "subcontractor company"],
    "category": ["category", "certification", "certification category", "description"],
}

CERTIFICATION_IMPORT_OPTIONAL_HEADERS = {
    "date_received": ["date received", "received", "start date", "date submitted"],
    "date_expires": ["date expires", "expiration", "expiration date", "expires"],
    "job": ["job", "job number"],
    "note": ["note", "notes"],
    "task": ["task", "pending task", "task description", "action required"],
    "task_note": ["task note", "task notes", "pending task note"],
}

CERTIFICATION_IMPORT_CUSTOM_ATTRIBUTE_HEADERS = {
    "Sponsor": ["sponsor"],
    "Date Submitted to Sponsor": ["date submitted to sponsor", "date submitted", "submitted to sponsor"],
    "Date Approved by Sponsor": ["date approved by sponsor", "date approved", "approved by sponsor"],
    "Access Finalized": ["access finalized", "access final"],
}

CERTIFICATION_IMPORT_CUSTOM_ATTRIBUTE_ALIASES = {
    "sponsor": "Sponsor",
    "date_submitted": "Date Submitted to Sponsor",
    "date_approved": "Date Approved by Sponsor",
    "access_finalized": "Access Finalized",
}


def _clean_import_value(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def _normalize_import_text(value):
    return re.sub(r"\s+", " ", _clean_import_value(value)).lower()


def _looks_like_certification_import_header(row_values):
    normalized = [_normalize_import_text(value) for value in row_values]
    known_headers = [
        header
        for headers in (
            list(CERTIFICATION_IMPORT_REQUIRED_HEADERS.values()) +
            list(CERTIFICATION_IMPORT_OPTIONAL_HEADERS.values())
        )
        for header in headers
    ]
    return any(value in known_headers for value in normalized if value)


def _certification_import_column_indexes(header_row):
    normalized_headers = [_normalize_import_text(value) for value in header_row]
    column_indexes = {"employee": 0, "subcontractor": 1, "category": 2}

    for field, possible_headers in {
        **CERTIFICATION_IMPORT_REQUIRED_HEADERS,
        **CERTIFICATION_IMPORT_OPTIONAL_HEADERS,
    }.items():
        for possible_header in possible_headers:
            if possible_header in normalized_headers:
                column_indexes[field] = normalized_headers.index(possible_header)
                break
    return column_indexes


def _certification_import_custom_attribute_column_indexes(header_row):
    normalized_headers = [_normalize_import_text(value) for value in header_row]
    column_indexes = {}
    for custom_attribute, possible_headers in CERTIFICATION_IMPORT_CUSTOM_ATTRIBUTE_HEADERS.items():
        for possible_header in possible_headers:
            if possible_header in normalized_headers:
                column_indexes[custom_attribute] = normalized_headers.index(possible_header)
                break
    return column_indexes


def _certification_import_cell(row_values, column_indexes, field):
    column_index = column_indexes.get(field)
    if column_index is None or column_index >= len(row_values):
        return ""
    return row_values[column_index]


def _certification_import_custom_attribute_values(row_values, custom_attribute_column_indexes):
    custom_attribute_values = {}
    for custom_attribute, column_index in custom_attribute_column_indexes.items():
        if column_index >= len(row_values):
            continue
        value = _clean_import_value(row_values[column_index])
        if value:
            custom_attribute_values[custom_attribute] = value
    return custom_attribute_values


def _parse_certification_import_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    cleaned_value = _clean_import_value(value)
    if not cleaned_value:
        return None
    try:
        parsed_value = parse_date(cleaned_value)
        if isinstance(parsed_value, datetime.datetime):
            return parsed_value.date()
        return parsed_value
    except (ValueError, TypeError, OverflowError):
        return None


def _employee_import_possible_names(employee):
    return {
        _normalize_import_text(f"{employee.first_name or ''} {employee.last_name or ''}"),
        _normalize_import_text(f"{employee.first_name or ''} {employee.middle_name or ''} {employee.last_name or ''}"),
        _normalize_import_text(f"{employee.last_name or ''}, {employee.first_name or ''}"),
        _normalize_import_text(str(employee)),
    }


def _find_certification_import_employee_matches(employee_name):
    normalized_name = _normalize_import_text(employee_name)
    if not normalized_name:
        return []

    matches = []
    employees = Employees.objects.filter(active=True)
    for employee in employees:
        if normalized_name in _employee_import_possible_names(employee):
            matches.append(employee)
    return matches


def _select_certification_import_employee_match(matches):
    if not matches:
        return None, []
    active_matches = [employee for employee in matches if employee.active]
    if len(active_matches) == 1:
        return active_matches[0], []
    if len(active_matches) > 1:
        return None, active_matches
    if len(matches) == 1:
        return matches[0], []
    return None, matches


def _find_certification_import_subcontractor(subcontractor_name):
    subcontractor_name = _clean_import_value(subcontractor_name)
    if not subcontractor_name:
        return None
    return Subcontractors.objects.filter(
        company__iexact=subcontractor_name,
        is_inactive=False,
    ).first()


def _find_certification_import_sub_employee(subcontractor, employee_name):
    employee_name = _clean_import_value(employee_name)
    if not subcontractor or not employee_name:
        return None
    return Subcontractor_Employees.objects.filter(
        subcontractor=subcontractor,
        name__iexact=employee_name,
        is_active=True,
    ).first()


def _find_certification_import_category(category_description):
    category_description = _clean_import_value(category_description)
    if not category_description:
        return None
    return CertificationCategories.objects.filter(description__iexact=category_description).select_related("template").first()


def _find_certification_import_job(job_number):
    job_number = _clean_import_value(job_number)
    if not job_number:
        return None
    return Jobs.objects.filter(job_number=job_number).first()


def _certification_import_rows(uploaded_file):
    if (getattr(uploaded_file, "name", "") or "").lower().endswith(".csv"):
        return _certification_import_csv_rows(uploaded_file)

    workbook = load_workbook(uploaded_file, data_only=True)
    worksheet = workbook.active
    raw_rows = [
        [cell for cell in row]
        for row in worksheet.iter_rows(values_only=True)
        if any(_clean_import_value(cell) for cell in row)
    ]
    if not raw_rows:
        return []

    first_row = raw_rows[0]
    has_header = _looks_like_certification_import_header(first_row)
    column_indexes = _certification_import_column_indexes(first_row) if has_header else {
        "employee": 0,
        "subcontractor": 1,
        "category": 2,
    }
    custom_attribute_column_indexes = (
        _certification_import_custom_attribute_column_indexes(first_row)
        if has_header
        else {}
    )
    start_row_number = 2 if has_header else 1
    rows = []

    for offset, row_values in enumerate(raw_rows[1:] if has_header else raw_rows, start=start_row_number):
        employee_name = _clean_import_value(_certification_import_cell(row_values, column_indexes, "employee"))
        subcontractor_name = _clean_import_value(_certification_import_cell(row_values, column_indexes, "subcontractor"))
        category_description = _clean_import_value(_certification_import_cell(row_values, column_indexes, "category"))
        if not employee_name and not subcontractor_name and not category_description:
            continue
        rows.append({
            "row_number": offset,
            "employee_name": employee_name,
            "subcontractor_name": subcontractor_name,
            "category_description": category_description,
            "date_received": _parse_certification_import_date(
                _certification_import_cell(row_values, column_indexes, "date_received")
            ),
            "date_expires": _parse_certification_import_date(
                _certification_import_cell(row_values, column_indexes, "date_expires")
            ),
            "job_number": _clean_import_value(_certification_import_cell(row_values, column_indexes, "job")),
            "note": _clean_import_value(_certification_import_cell(row_values, column_indexes, "note")),
            "task": _clean_import_value(_certification_import_cell(row_values, column_indexes, "task")),
            "task_note": _clean_import_value(_certification_import_cell(row_values, column_indexes, "task_note")),
            "custom_attribute_values": _certification_import_custom_attribute_values(
                row_values,
                custom_attribute_column_indexes,
            ),
        })
    return rows


def _certification_import_csv_rows(uploaded_file):
    raw_content = uploaded_file.read()
    if isinstance(raw_content, str):
        text_content = raw_content
    else:
        text_content = raw_content.decode("utf-8-sig")
    raw_rows = [
        row for row in csv.reader(text_content.splitlines())
        if any(_clean_import_value(cell) for cell in row)
    ]
    if not raw_rows:
        return []

    first_row = raw_rows[0]
    has_header = _looks_like_certification_import_header(first_row)
    column_indexes = _certification_import_column_indexes(first_row) if has_header else {
        "employee": 0,
        "subcontractor": 1,
        "category": 2,
    }
    custom_attribute_column_indexes = (
        _certification_import_custom_attribute_column_indexes(first_row)
        if has_header
        else {}
    )
    start_row_number = 2 if has_header else 1
    rows = []

    for offset, row_values in enumerate(raw_rows[1:] if has_header else raw_rows, start=start_row_number):
        employee_name = _clean_import_value(_certification_import_cell(row_values, column_indexes, "employee"))
        subcontractor_name = _clean_import_value(_certification_import_cell(row_values, column_indexes, "subcontractor"))
        category_description = _clean_import_value(_certification_import_cell(row_values, column_indexes, "category"))
        if not employee_name and not subcontractor_name and not category_description:
            continue
        rows.append({
            "row_number": offset,
            "employee_name": employee_name,
            "subcontractor_name": subcontractor_name,
            "category_description": category_description,
            "date_received": _parse_certification_import_date(
                _certification_import_cell(row_values, column_indexes, "date_received")
            ),
            "date_expires": _parse_certification_import_date(
                _certification_import_cell(row_values, column_indexes, "date_expires")
            ),
            "job_number": _clean_import_value(_certification_import_cell(row_values, column_indexes, "job")),
            "note": _clean_import_value(_certification_import_cell(row_values, column_indexes, "note")),
            "task": _clean_import_value(_certification_import_cell(row_values, column_indexes, "task")),
            "task_note": _clean_import_value(_certification_import_cell(row_values, column_indexes, "task_note")),
            "custom_attribute_values": _certification_import_custom_attribute_values(
                row_values,
                custom_attribute_column_indexes,
            ),
        })
    return rows


def _normalize_certification_import_custom_attribute_name(custom_attribute):
    return _normalize_import_text(custom_attribute).replace(" ", "_")


def _apply_certification_import_custom_attribute_values(certification, custom_attribute_values):
    updated_count = 0
    if not custom_attribute_values:
        return updated_count

    custom_attributes = CertificationCustomAttributes.objects.filter(certification=certification)
    custom_attributes_by_name = {
        _normalize_certification_import_custom_attribute_name(attribute.custom_attribute): attribute
        for attribute in custom_attributes
    }

    for custom_attribute_name, custom_attribute_value in custom_attribute_values.items():
        custom_attribute_name = CERTIFICATION_IMPORT_CUSTOM_ATTRIBUTE_ALIASES.get(
            custom_attribute_name,
            custom_attribute_name,
        )
        custom_attribute = custom_attributes_by_name.get(
            _normalize_certification_import_custom_attribute_name(custom_attribute_name)
        )
        if not custom_attribute:
            continue
        custom_attribute.custom_attribute_result = custom_attribute_value
        custom_attribute.save(update_fields=["custom_attribute_result"])
        updated_count += 1
    return updated_count


def _create_import_pending_task(certification, current_employee, task_description, task_note):
    if not task_description:
        return {"created": False, "email_sent": False, "warning": ""}

    task_employee = certification.employee
    task_subcontractor_employee = certification.subcontractor_employee
    recipient_email = (
        task_subcontractor_employee.email
        if task_subcontractor_employee
        else task_employee.email if task_employee else ""
    )

    notes = ""
    if task_note:
        note_prefix = date.today().strftime("%m/%d/%Y")
        if current_employee:
            note_prefix += f" - {current_employee.first_name}"
        notes = f"{note_prefix}: {task_note}"

    pending_action = EmployeePendingActions.objects.create(
        employee=task_employee,
        subcontractor_employee=task_subcontractor_employee,
        certification=certification,
        date=date.today(),
        description=task_description,
        notes=notes,
        is_complete=False,
        confirmed_is_complete=False,
    )
    assignee_display = (
        f"[SUB] {task_subcontractor_employee}"
        if task_subcontractor_employee
        else str(task_employee)
    )
    certification_note = f"Pending employee task assigned to {assignee_display}: {task_description}"
    if task_note:
        certification_note += f". Note: {task_note}"
    CertificationNotes.objects.create(
        certification=certification,
        date=date.today(),
        user=current_employee or task_employee or Employees.objects.filter(user__is_superuser=True).first() or Employees.objects.first(),
        note=certification_note,
    )

    if not recipient_email:
        return {
            "created": True,
            "email_sent": False,
            "warning": f"Row task created for {assignee_display}, but no email address was found.",
        }

    sender = (
        current_employee.email
        if current_employee and current_employee.email
        else "bridgette@gerloffpainting.com"
    )
    email_body = (
        "A new required task has been added for you.\n\n"
        f"Task: {pending_action.description}\n"
        f"Certification: {certification}\n"
        f"Date Added: {pending_action.date.strftime('%m/%d/%Y')}\n"
    )
    if task_note:
        email_body += f"\nNotes:\n{task_note}"
    try:
        Email.sendEmail("New Required Task", email_body, [recipient_email], False, sender)
    except Exception:
        return {
            "created": True,
            "email_sent": False,
            "warning": f"Row task created for {assignee_display}, but the email could not be sent.",
        }
    return {"created": True, "email_sent": True, "warning": ""}


def _review_certification_import_rows(rows):
    reviewed_rows = []
    errors = []
    warnings = []

    for row in rows:
        row_errors = []
        row_warnings = []
        employee = None
        subcontractor = None
        subcontractor_employee = None
        would_create_subcontractor_employee = False

        if not row["employee_name"]:
            row_errors.append("Missing employee name.")

        if not row["category_description"]:
            row_errors.append("Missing certification/category description.")

        if row["subcontractor_name"]:
            subcontractor = _find_certification_import_subcontractor(row["subcontractor_name"])
            if not subcontractor:
                row_errors.append(f"Subcontractor not found: {row['subcontractor_name']}.")
            else:
                subcontractor_employee = _find_certification_import_sub_employee(subcontractor, row["employee_name"])
                if not subcontractor_employee and row["employee_name"]:
                    would_create_subcontractor_employee = True
        else:
            employee_matches = _find_certification_import_employee_matches(row["employee_name"])
            employee, ambiguous_employee_matches = _select_certification_import_employee_match(employee_matches)
            if ambiguous_employee_matches:
                match_list = ", ".join(
                    f"{match} (ID {match.id})"
                    for match in ambiguous_employee_matches
                )
                row_errors.append(f"Multiple employees found for {row['employee_name']}: {match_list}.")
            elif not employee and row["employee_name"]:
                row_errors.append(f"Employee not found: {row['employee_name']}.")

        category = _find_certification_import_category(row["category_description"])
        if row["category_description"] and not category:
            row_warnings.append(
                f"Category not found: {row['category_description']}. Final upload will use this as the certification description without linking a category."
            )

        job = _find_certification_import_job(row["job_number"])
        if row["job_number"] and not job:
            row_warnings.append(f"Job not found: {row['job_number']}. Certification will import without a job.")

        reviewed_row = {
            **row,
            "employee": employee,
            "subcontractor": subcontractor,
            "subcontractor_employee": subcontractor_employee,
            "would_create_subcontractor_employee": would_create_subcontractor_employee,
            "category": category,
            "job": job,
            "errors": row_errors,
            "warnings": row_warnings,
        }
        reviewed_rows.append(reviewed_row)
        for error in row_errors:
            errors.append(f"Row {row['row_number']}: {error}")
        for warning in row_warnings:
            warnings.append(f"Row {row['row_number']}: {warning}")
    return reviewed_rows, errors, warnings


def _commit_certification_import(reviewed_rows, request_user):
    current_employee = Employees.objects.filter(user=request_user).first()
    results = {
        "certifications_created": 0,
        "subcontractor_employees_created": 0,
        "custom_attributes_created": 0,
        "custom_attributes_updated": 0,
        "pending_tasks_created": 0,
        "task_emails_sent": 0,
        "warnings": [],
        "created_rows": [],
    }

    with transaction.atomic():
        for row in reviewed_rows:
            subcontractor_employee = row["subcontractor_employee"]
            if row["subcontractor"] and not subcontractor_employee:
                subcontractor_employee = Subcontractor_Employees.objects.create(
                    subcontractor=row["subcontractor"],
                    name=row["employee_name"],
                    date_enrolled=date.today(),
                    is_active=True,
                )
                results["subcontractor_employees_created"] += 1

            certification = Certifications.objects.create(
                category=row["category"],
                employee=row["employee"],
                subcontractor=row["subcontractor"],
                subcontractor_employee=subcontractor_employee,
                description=row["category"].description if row["category"] else row["category_description"],
                date_received=row["date_received"],
                date_expires=row["date_expires"],
                job=row["job"],
                note=row["note"],
            )
            results["certifications_created"] += 1
            results["custom_attributes_created"] += _create_standard_certification_custom_attributes(certification)
            results["custom_attributes_updated"] += _apply_certification_import_custom_attribute_values(
                certification,
                row["custom_attribute_values"],
            )

            if row["note"]:
                CertificationNotes.objects.create(
                    certification=certification,
                    date=date.today(),
                    user=current_employee or row["employee"] or Employees.objects.filter(user__is_superuser=True).first() or Employees.objects.first(),
                    note=row["note"],
                )

            task_result = _create_import_pending_task(
                certification,
                current_employee,
                row["task"],
                row["task_note"],
            )
            if task_result["created"]:
                results["pending_tasks_created"] += 1
            if task_result["email_sent"]:
                results["task_emails_sent"] += 1
            if task_result["warning"]:
                results["warnings"].append(f"Row {row['row_number']}: {task_result['warning']}")

            results["created_rows"].append({
                "row_number": row["row_number"],
                "certification_id": certification.id,
                "description": certification.description,
                "employee_display": (
                    f"[SUB] {subcontractor_employee}"
                    if subcontractor_employee
                    else str(row["employee"])
                ),
            })

    return results


def _has_employee_toolbox_record(toolbox_talk, employee):
    return CompletedToolboxTalks.objects.filter(
        master=toolbox_talk,
        employee=employee
    ).exists()


def _has_sub_employee_toolbox_record(toolbox_talk, employee, job):
    if CompletedSubToolboxTalks.objects.filter(
        master=toolbox_talk,
        employee=employee,
        job=job
    ).exists():
        return True

    return CompletedSubToolboxJobTalkEmployees.objects.filter(
        employee=employee,
        completed__scheduled=toolbox_talk,
        completed__job=job,
        completed__is_excused=False
    ).exists()


def _has_sub_job_toolbox_record(toolbox_talk, subcontractor, job):
    return CompletedSubToolboxJobTalks.objects.filter(
        scheduled=toolbox_talk,
        subcontractor=subcontractor,
        job=job
    ).exists()


def _pending_employee_task_queryset():
    return (
        EmployeePendingActions.objects
        .filter(is_complete=False)
        .filter(
            Q(employee__active=True) |
            Q(
                subcontractor_employee__is_active=True,
                subcontractor_employee__subcontractor__is_inactive=False,
            )
        )
        .select_related(
            "employee",
            "employee__job_title",
            "subcontractor_employee",
            "subcontractor_employee__subcontractor",
            "certification",
            "certification__category",
        )
        .order_by(
            "employee__last_name",
            "employee__first_name",
            "subcontractor_employee__subcontractor__company",
            "subcontractor_employee__name",
            "date",
            "id",
        )
    )


def _certification_options_for_pending_task(task):
    if task.employee_id:
        certifications = Certifications.objects.filter(
            employee=task.employee,
            is_closed=False,
        )
    elif task.subcontractor_employee_id:
        certifications = Certifications.objects.filter(
            subcontractor_employee=task.subcontractor_employee,
            is_closed=False,
        )
    else:
        certifications = Certifications.objects.none()

    certifications = list(
        certifications
        .select_related("category", "job")
        .order_by("category__description", "description", "-date_received", "id")
    )

    if task.certification and task.certification_id not in {cert.id for cert in certifications}:
        certifications.append(task.certification)

    _apply_certification_display_descriptions(certifications)
    if task.certification:
        for certification in certifications:
            if certification.id == task.certification_id:
                task.certification.display_description = certification.display_description
                break
    return certifications


def _set_pending_task_certification(pending_task, certification_id):
    if certification_id:
        certification = get_object_or_404(
            Certifications.objects.select_related("category"),
            id=certification_id,
        )
        if certification.is_closed and certification.id != pending_task.certification_id:
            return "That certification is closed and cannot be linked to this task."
        if pending_task.employee_id and certification.employee_id != pending_task.employee_id:
            return "That certification is not linked to this employee."
        if (
            pending_task.subcontractor_employee_id and
            certification.subcontractor_employee_id != pending_task.subcontractor_employee_id
        ):
            return "That certification is not linked to this subcontractor employee."
        pending_task.certification = certification
    else:
        pending_task.certification = None
    return ""


def _expired_flagged_certification_queryset():
    return (
        Certifications.objects
        .filter(
            is_closed=False,
            is_flagged_when_expired=True,
            date_expires__lt=date.today(),
        )
        .filter(
            Q(employee__active=True) |
            Q(
                employee__isnull=True,
                subcontractor__is_inactive=False,
                subcontractor_employee__isnull=True,
            ) |
            Q(
                employee__isnull=True,
                subcontractor__is_inactive=False,
                subcontractor_employee__is_active=True,
            )
        )
        .select_related(
            "employee",
            "category",
            "job",
            "subcontractor",
            "subcontractor_employee",
        )
        .order_by("date_expires", "employee__last_name", "employee__first_name", "subcontractor__company")
    )


def _get_first_subcontractor_invoice_date(subcontract):
    return (
        SubcontractorInvoice.objects
        .filter(subcontract=subcontract)
        .order_by('date', 'id')
        .values_list('date', flat=True)
        .first()
    )


def _get_subcontractor_employee_delegation(subcontractor, subcontract):
    return SubcontractorEmployeeDelegation.objects.filter(
        subcontractor=subcontractor,
        subcontract=subcontract
    ).first()


def _subcontractor_employee_delegation_effective(subcontractor, subcontract, scheduled_date=None):
    delegation = _get_subcontractor_employee_delegation(subcontractor, subcontract)
    return bool(delegation)


def _subcontract_has_delegated_employee_for_date(subcontract, scheduled_date):
    if not _subcontractor_employee_delegation_effective(
        subcontract.subcontractor,
        subcontract,
        scheduled_date
    ):
        return False

    return Subcontractor_Job_Assignments.objects.filter(
        job=subcontract.job_number,
        employee__subcontractor=subcontract.subcontractor,
        employee__is_active=True,
        employee__has_access_to_toolbox=True,
        employee__date_enrolled__isnull=False,
        employee__date_enrolled__lte=scheduled_date
    ).exists()


def _sub_employee_has_active_toolbox_job(sub_employee, job=None):
    query = Subcontractor_Job_Assignments.objects.filter(
        employee=sub_employee,
        job__is_closed=False,
        job__is_active=True,
        job__is_labor_done=False
    )

    if job is not None:
        query = query.filter(job=job)

    return query.exists()


def _get_delegated_active_toolbox_jobs_for_sub_employee(sub_employee, job=None, effective_date=None):
    if (
        not sub_employee.has_access_to_toolbox or
        not sub_employee.subcontractor or
        not sub_employee.subcontractor.is_toolbox_required
    ):
        return []

    assignments = Subcontractor_Job_Assignments.objects.filter(
        employee=sub_employee,
        job__is_closed=False,
        job__is_active=True,
        job__is_labor_done=False
    ).select_related('job')

    if job is not None:
        assignments = assignments.filter(job=job)

    delegated_jobs = []
    seen = set()
    for assignment in assignments:
        subcontract = Subcontracts.objects.filter(
            subcontractor=sub_employee.subcontractor,
            job_number=assignment.job,
            is_closed=False,
            subcontractor__is_toolbox_required=True,
            job_number__is_closed=False,
            job_number__is_active=True,
            job_number__is_labor_done=False
        ).first()

        if not subcontract:
            continue

        if not _subcontractor_employee_delegation_effective(
            subcontractor=sub_employee.subcontractor,
            subcontract=subcontract,
            scheduled_date=effective_date
        ):
            continue

        if assignment.job_id in seen:
            continue
        seen.add(assignment.job_id)
        delegated_jobs.append(assignment.job)

    return delegated_jobs


def _has_sub_employee_toolbox_record_for_any_job(sub_employee, toolbox_talk, jobs):
    if CompletedSubToolboxTalks.objects.filter(
        master=toolbox_talk,
        employee=sub_employee,
        job__in=jobs
    ).exists():
        return True

    return CompletedSubToolboxJobTalkEmployees.objects.filter(
        employee=sub_employee,
        completed__scheduled=toolbox_talk,
        completed__job__in=jobs,
        completed__is_excused=False
    ).exists()


def _count_missing_toolbox_talks_before(cutoff_date):
    missing_toolbox_talks = 0
    employee_titles = ["Painter", "Warehouse", "Superintendent"]

    for toolbox_talk in ScheduledToolboxTalks.objects.filter(
        date__lt=cutoff_date
    ).order_by('-date'):
        counted_employees = set()

        employee_assignments = ScheduledToolboxTalkEmployees.objects.filter(
            scheduled=toolbox_talk,
            employee__active=True,
            employee__job_title__description__in=employee_titles
        ).select_related(
            'employee'
        )

        for assignment in employee_assignments:
            employee_key = assignment.employee_id
            if employee_key in counted_employees:
                continue
            counted_employees.add(employee_key)

            if not _has_employee_toolbox_record(toolbox_talk, assignment.employee):
                missing_toolbox_talks += 1

        if not toolbox_talk.is_all_employees:
            continue

        target_employees = Employees.objects.filter(
            active=True,
            date_added__lte=toolbox_talk.date,
            job_title__description__in=employee_titles
        )

        for employee in target_employees:
            employee_key = employee.id
            if employee_key in counted_employees:
                continue
            counted_employees.add(employee_key)

            if not _has_employee_toolbox_record(toolbox_talk, employee):
                missing_toolbox_talks += 1

    missing_toolbox_talks += sub_toolbox.get_missing_subcontractor_count_before(cutoff_date)

    return missing_toolbox_talks




@login_required(login_url='/accounts/login')
def seperate_test(request):
    fileitem = request.FILES['filename']
    fn = os.path.basename(fileitem.name)
    fn2 = os.path.join("C:/Trinity/", fn)
    open(fn2, 'wb').write(fileitem.file.read())
    return redirect('index')


@login_required(login_url='/accounts/login')
def client_info_job(request, jobnumber):
    send_data = {}
    job = Jobs.objects.get(job_number=jobnumber)
    client = job.client
    employees = ClientEmployees.objects.filter(id=client,is_active=True).order_by('name')
    send_data['job'] =job
    send_data['client'] =client
    send_data['employees'] =employees
    clientemployees = []
    for person in employees:
        changeorder= False
        for x in ClientJobRoles.objects.filter(job=job):


            clientemployees.append({'person':'test'})


@login_required(login_url='/accounts/login')
def client_job_info(request, id):
    send_data = {}
    selected_job = Jobs.objects.get(job_number=id)
    selected_client = selected_job.client
    send_data['selected_client'] = selected_client
    send_data['selected_job'] = selected_job
    send_data['error_message'] = ""
    if request.method == "POST":
        if 'go_back' in request.POST:
            return redirect('client_info', id=selected_client.id)
        selected_employee = ClientEmployees.objects.get(person_pk=request.POST['person_pk'])
        person_pk = request.POST['person_pk']
        if 'pm' + person_pk in request.POST:
            if selected_job.client_Pm != selected_employee:
                if selected_employee.email:
                    selected_job.client_Pm = selected_employee
                    selected_job.save()
                else:
                    send_data['error_message'] = "Employee must have an email address in order to be PM"
        if 'super' + person_pk in request.POST:
            selected_job.client_Super = selected_employee
            selected_job.save()
        if 'changeorder' + person_pk in request.POST:
            if not ClientJobRoles.objects.filter(job=selected_job, employee=selected_employee,
                                             role="Change Orders").exists():
                ClientJobRoles.objects.create(job=selected_job, employee=selected_employee,
                                             role="Change Orders")
        else:
            # if ClientJobRoles.objects.filter(job=selected_job, employee=selected_employee,
            #                                  role="Change Orders").exists():
            ClientJobRoles.objects.filter(job=selected_job, employee=selected_employee,
                                              role="Change Orders").delete()
        if 'ewt' + person_pk in request.POST:
            if not ClientJobRoles.objects.filter(job=selected_job, employee=selected_employee,
                                             role="Extra Work Tickets").exists():
                ClientJobRoles.objects.create(job=selected_job, employee=selected_employee,
                                             role="Extra Work Tickets")
        else:
            ClientJobRoles.objects.filter(job=selected_job, employee=selected_employee,
                                              role="Extra Work Tickets").delete()
    people=[]
    for x in ClientEmployees.objects.filter(id=selected_client,is_active=True).order_by('name'):
        changeorder='No'
        ewt='No'
        super ='No'
        pm='No'
        if selected_job.client_Super == x: super = 'Yes'
        if selected_job.client_Pm == x: pm = 'Yes'
        if ClientJobRoles.objects.filter(job=selected_job, employee=x,role="Change Orders").exists(): changeorder='Yes'
        if ClientJobRoles.objects.filter(job=selected_job, employee=x,
                                         role="Extra Work Tickets").exists(): ewt = 'Yes'
        people.append({'person':x,'changeorder': changeorder, 'ewt':ewt,'super':super,'pm':pm})
    send_data['people'] = people

    return render(request, 'client_job_info.html', send_data)

@login_required(login_url='/accounts/login')
def client_info(request, id):
    send_data = {}
    send_data['clients']=Clients.objects.filter(is_active=True).order_by('company')
    send_data['active_clients']=Clients.objects.filter(is_active=True).order_by('company')
    send_data['all_clients']=Clients.objects.all().order_by('company')
    if id != 'ALL':
        selected_client = Clients.objects.get(id=id)
        send_data['selected_client'] = selected_client
        send_data['client_employees'] = ClientEmployees.objects.filter(id=selected_client,is_active=True).order_by('name')
        send_data['jobs']=Jobs.objects.filter(client=selected_client, is_closed=False)
    if request.method == "POST":
        if 'combine_companies_now' in request.POST:
            company1 = Clients.objects.get(id=request.POST['select_client1'])
            company2 = Clients.objects.get(id=request.POST['select_client2'], is_active=True)
            for x in ClientEmployees.objects.filter(id=company1,is_active=True).order_by('name'):
                x.id=company2
                x.save()
            for x in Jobs.objects.filter(client=company1):
                x.client=company2
                x.save()
            company1.delete()
            return redirect('client_info',id=company2.id)
        if 'search_client' in request.POST:
            send_data['clients'] = Clients.objects.filter(company__icontains=request.POST['search_client'], is_active=True).order_by('company')
            send_data['search_client_word'] = request.POST['search_client']
        if 'search_job' in request.POST:
            send_data['jobs'] = Jobs.objects.filter(client=selected_client, job_name__icontains=request.POST['search_job'])
            send_data['search_job_word'] = request.POST['search_job']
        if 'select_client' in request.POST:
            if request.POST['select_client'] != 'please_select':
                return redirect('client_info', id=request.POST['select_client'])
                # selected_client = Clients.objects.get(id=request.POST['select_client'])
                # send_data['selected_client'] = selected_client
        if 'select_job' in request.POST:
            if request.POST['select_job'] != 'please_select':
                return redirect('client_job_info', id=request.POST['select_job'])
        if 'make_client_inactive' in request.POST:
            selected_client.is_active=False
            selected_client.save()
        if 'make_client_active' in request.POST:
            selected_client.is_active=True
            selected_client.save()
        if 'company_name' in request.POST:
            selected_client.company = request.POST['company_name']
            selected_client.bid_email = request.POST['company_email']
            selected_client.phone = request.POST['company_phone']
            selected_client.address = request.POST['company_address']
            selected_client.city = request.POST['company_city']
            selected_client.state = request.POST['company_state']
            selected_client.save()
        if 'combine_people_now' in request.POST:
            person1 = ClientEmployees.objects.get(person_pk=request.POST['select_person1'])
            person2 = ClientEmployees.objects.get(person_pk=request.POST['select_person2'])

            Jobs.objects.filter(client_Pm=person1).update(client_Pm=person2)
            Jobs.objects.filter(client_Super=person1).update(client_Super=person2)

            ClientJobRoles.objects.filter(employee=person1).update(employee=person2)
            TempRecipients.objects.filter(person=person1).update(person=person2)
            TempRecipientsCOPList.objects.filter(person=person1).update(person=person2)

            person1.delete()
        if 'people_form' in request.POST:
            for x in request.POST:
                if x[0:4]=='name':
                    current_person_pk = x[4:len(x)]
                    current_person = ClientEmployees.objects.get(person_pk=current_person_pk)
                    current_person.name = request.POST[x]
                    current_person.email = request.POST['email' + current_person_pk]
                    current_person.phone = request.POST['phone' + current_person_pk]
                    current_person.save()
                    if 'closed' + current_person_pk in request.POST:
                        if current_person.is_active:
                            job_pm = Jobs.objects.filter(client_Pm=current_person,is_closed=False).values_list('job_name', flat=True)
                            jobrole = ClientJobRoles.objects.filter(employee=current_person,job__is_closed=True).values_list('job__job_name', flat=True)
                            job_super = Jobs.objects.filter(client_Super=current_person,is_closed=False).values_list('job_name', flat=True)
                            job_pm_list = ",".join(job_pm)
                            jobrole_list = ",".join(jobrole)
                            job_super_list = ",".join(job_super)
                            if job_pm:
                                messages.error(request, f"This person is assigned as PM to {job_pm_list}. You must change the PM before marking inactive")
                            if jobrole_list:
                                messages.error(request,f"This person is the default CO recipient for {jobrole_list}. You must change this first before marking inactive")
                            if job_super:
                                messages.error(request,
                                               f"This person is assigned as Super to {job_super_list}. You must change the Super before marking inactive")
                        # try:
                        #     Email.sendEmail("Respirator Clearance Completed", message,
                        #                     recipients, False)
                        #     message = "Your email about the respirator clearance was sent successfully"
                        #     messages.error(request,
                        #                    "There was a problem sending the email to Bridgette. Please tell her it is approved.")
                        # except:
                        #     message = "Error! Your email about the respirator clearance failed to send. Please call them and let them know it was completed."
                            if not job_pm and not job_super and not jobrole:
                                current_person.is_active=False
                    else:
                        current_person.is_active = True
                    current_person.save()
            if 'add_new_person' in request.POST:
                ClientEmployees.objects.create(id=selected_client, name=request.POST['add_name'], phone=request.POST['add_phone'], email = request.POST['add_email'])
            send_data['client_employees'] = ClientEmployees.objects.filter(id=selected_client,is_active=True).order_by('name')

    return render(request, 'client_info.html', send_data)


@login_required(login_url='/accounts/login')
def index(request):
    current_employee = Employees.objects.get(user=request.user)
    if current_employee.job_title.description == "Painter":
        return redirect('my_page')
    if current_employee.job_title.description == "Superintendent":
        return redirect('super_home', super='AUTO')

    if current_employee.job_title.description == "Warehouse":
        return redirect('warehouse_home')
    send_data = {}
    next_two_weeks = 0
    for x in Jobs.objects.filter(is_closed=False, is_active=False, is_labor_done=False):
        if x.next_two_weeks() == True:
            next_two_weeks += 1
    send_data['needs_work_order_now'] = Jobs.objects.filter(is_work_order_done=False, is_closed=False,start_date__lt=date.today()+ timedelta(days=21)).count()
    send_data['needs_work_order'] = Jobs.objects.filter(is_work_order_done=False, is_closed=False,).count()
    last_clockshark_import = (
        ClockSharkTimeEntry.objects
        .filter(hours_adjust_note="AUTO IMPORT")
        .aggregate(last_import=Max("work_day"))
    )["last_import"]

    send_data["last_clockshark_import"] = last_clockshark_import
    end_date = timezone.localdate()
    start_date = subtract_months(end_date, 4)

    unmatched_jobs_count = (
        ClockSharkTimeEntry.objects
        .filter(
            job__isnull=True,
            work_day__gte=start_date,
            work_day__lte=end_date,
        )
        .exclude(job_name__isnull=True)
        .exclude(job_name="")
        .exclude(job_name="Gerloff Painting Inc")
        .exclude(job_name="Sick Leave")
        .exclude(job_name="Requested Day Off")
        .values("job_name")
        .distinct()
        .count()
    )

    send_data["unmatched_jobs_count"] = unmatched_jobs_count

    send_data['missing'] = Inventory.objects.filter(status="Missing", is_closed=False).count()
    send_data['checked_out'] = Inventory.objects.filter(job_number__is_closed=False, is_closed=False).count()
    send_data['closed_job'] = Inventory.objects.filter(job_number__is_closed=True, is_closed=False).count()
    send_data['service'] = Inventory.objects.filter(service_vendor__isnull=False, is_closed=False).count()
    send_data['pickup_requests'] = PickupRequest.objects.filter(is_closed=False).count()
    send_data['rentals'] = Rentals.objects.filter(off_rent_number=None, is_closed=False).count()
    check_rentals=0
    for x in Rentals.objects.filter(off_rent_number=None, is_closed=False):
        if x.colorize(): check_rentals += 1
    send_data['check_rentals'] = check_rentals
    send_data['rentals_requested_off'] = Rentals.objects.filter(requested_off_rent=True, off_rent_number=None, is_closed=False).count()
    send_data['next_two_weeks'] = next_two_weeks
    send_data['needs_super'] = Jobs.objects.filter(superintendent__isnull=True, is_closed=False).count()
    send_data['active_subcontracts'] = Subcontracts.objects.filter(job_number__is_closed=False, is_closed=False).count()
    send_data['pending_invoices'] = SubcontractorInvoice.objects.filter(is_sent=False).count()
    send_data['approved_invoices'] = SubcontractorInvoice.objects.filter(is_sent=True, processed=False).count()
    send_data['pending_employee_tasks_count'] = _pending_employee_task_queryset().count()
    send_data['expired_flagged_certifications_count'] = _expired_flagged_certification_queryset().count()
    send_data['need_to_be_closed'] = Jobs.objects.filter(is_labor_done=True,is_closed=False).count()
    send_data['unsuccessful_login_attempts_past_week'] = LoginAttempt.objects.filter(
        result=LoginAttempt.RESULT_FAILED,
        attempted_at__gte=timezone.now() - timedelta(days=7)
    ).count()
    send_data['unapproved_sub_changes'] = SubcontractItems.objects.filter(is_approved=False, subcontract__is_closed=False).count()
    if Jobs.objects.filter(is_closed=False, superintendent=Employees.objects.get(
            user=request.user)).exists() and request.user != Employees.objects.get(id=22).user:
        active_super = Employees.objects.get(user=request.user)
        send_data['super_equipment'] = Inventory.objects.filter(job_number__superintendent=active_super,
                                                                is_closed=False).count()  #
        send_data['super_rentals'] = Rentals.objects.filter(job_number__superintendent=active_super, is_closed=False,requested_off_rent=False).count()  #
        send_data['active_subcontracts'] = Subcontracts.objects.filter(job_number__superintendent=active_super,
                                                                       job_number__is_closed=False,
                                                                       is_closed=False).count()
        send_data['pending_invoices'] = SubcontractorInvoice.objects.filter(
            subcontract__job_number__superintendent=active_super, is_sent=False).count()
        send_data['tickets'] = 0  #
        send_data['active_jobs'] = Jobs.objects.filter(superintendent=active_super, is_active=True, is_closed=False).count()
        send_data['punchlist_jobs'] = Jobs.objects.filter(superintendent=active_super,
                                                          is_waiting_for_punchlist=True,is_closed=False).count()

        next_two_weeks = 0
        for x in Jobs.objects.filter(is_closed=False, is_active=False, superintendent=active_super, is_labor_done=False):
            if x.next_two_weeks() == True:
                next_two_weeks += 1
    else:
        send_data['super_equipment'] = Inventory.objects.filter(job_number__is_closed=False, is_closed=False).count()
        send_data['super_rentals'] = Rentals.objects.filter(off_rent_number=None, is_closed=False).count()
        send_data['tickets'] = 0  #
        send_data['active_jobs'] = Jobs.objects.filter(is_active=True,is_closed=False).count()
        send_data['punchlist_jobs'] = Jobs.objects.filter(is_waiting_for_punchlist=True,is_closed=False).count()
        next_two_weeks = 0
        for x in Jobs.objects.filter(is_closed=False, is_active=False, is_labor_done=False):
            if x.next_two_weeks() == True:
                next_two_weeks += 1
    send_data['super_jobs'] = next_two_weeks  #
    send_data['current_user'] = request.user.first_name
    clearance_forms_needing_review = len(get_respirators_in_review())
    send_data['clearance_forms_needing_review']=clearance_forms_needing_review
    painters_needing_respirator=0
    for x in Employees.objects.filter(job_title__description="Painter",active=True):
        if not RespiratorClearance.objects.filter(employee=x).exists():
            if not Certifications.objects.filter(employee=x,category__description="Respirator Clearance").exists():
                painters_needing_respirator+=1
    send_data['painters_needing_respirator'] = painters_needing_respirator
    today = date.today()
    send_data['missing_toolbox_talks'] = _count_missing_toolbox_talks_before(today + timedelta(days=1))
    wallcoverings = Wallcovering.objects.filter(
        job_number__is_closed=False,
        is_owner_furnished=False
    ).select_related(
        "job_number",
        "vendor"
    )

    wc_not_submitted_count = 0
    wc_not_approved_count = 0
    wc_approved_not_ordered_count = 0

    for wc in wallcoverings:
        submittal_status = wc.submittal_status()
        ordering_status = wc.ordering_status()

        has_approved_cop_not_ordered = Wallcovering_Change_Orders.objects.filter(
            wallcovering=wc,
            is_ordered=False,
            change_order__is_approved=True
        ).exists()

        # Submittal buckets
        if submittal_status in ["Not Submitted", "Partially Submitted"]:
            wc_not_submitted_count += 1

        elif submittal_status == "Submitted":
            wc_not_approved_count += 1

        # Ordering bucket - separate from submittal status
        if (
                (
                        submittal_status == "Approved"
                        and ordering_status == "Not Ordered"
                )
                or has_approved_cop_not_ordered
        ):
            wc_approved_not_ordered_count += 1

    send_data['wc_not_submitted_count'] =wc_not_submitted_count
    send_data['wc_not_approved_count'] =wc_not_approved_count
    send_data['wc_approved_not_ordered_count'] =wc_approved_not_ordered_count

    return render(request, 'index.html', send_data)


@login_required(login_url='/accounts/login')
def pending_employee_tasks(request):
    if request.method == "POST" and "add_pending_task_note" in request.POST:
        pending_task = get_object_or_404(
            _pending_employee_task_queryset(),
            id=request.POST.get("pending_task_id"),
        )
        note = (request.POST.get(f"pending_task_note_{pending_task.id}") or "").strip()
        certification_error = _set_pending_task_certification(
            pending_task,
            request.POST.get("certification_id"),
        )
        if certification_error:
            messages.error(request, certification_error)
            return redirect("pending_employee_tasks")
        if not note:
            messages.error(request, "Enter a note before submitting.")
            return redirect("pending_employee_tasks")

        current_employee = Employees.objects.filter(user=request.user).first()
        note_prefix = date.today().strftime('%m/%d/%Y')
        if current_employee:
            note_prefix += f" - {current_employee.first_name}"
        else:
            note_prefix += f" - {request.user.get_full_name() or request.user.username}"
        pending_task.notes = (
            (pending_task.notes + "\n") if pending_task.notes else ""
        ) + f"{note_prefix}: {note}"
        pending_task.save(update_fields=["certification", "notes"])
        messages.success(request, "Task note added.")
        return redirect("pending_employee_tasks")

    if request.method == "POST" and "delete_pending_task" in request.POST:
        pending_task = get_object_or_404(
            _pending_employee_task_queryset(),
            id=request.POST.get("pending_task_id"),
        )
        current_employee = Employees.objects.filter(user=request.user).first()
        if pending_task.certification:
            CertificationNotes.objects.create(
                certification=pending_task.certification,
                date=date.today(),
                user=current_employee or pending_task.employee or Employees.objects.filter(user__is_superuser=True).first() or Employees.objects.first(),
                note=(
                    f"Pending employee task deleted for {pending_task.assignee_display}: "
                    f"{pending_task.description}"
                ),
            )
        pending_task.delete()
        messages.success(request, "Pending employee task deleted.")
        return redirect("pending_employee_tasks")

    if request.method == "POST" and "complete_pending_task" in request.POST:
        pending_task = get_object_or_404(
            _pending_employee_task_queryset(),
            id=request.POST.get("pending_task_id"),
        )
        certification_id = request.POST.get("certification_id")
        completion_note = (request.POST.get("completion_note") or "").strip()
        current_employee = Employees.objects.filter(user=request.user).first()

        certification_error = _set_pending_task_certification(pending_task, certification_id)
        if certification_error:
            messages.error(request, certification_error)
            return redirect("pending_employee_tasks")

        completed_by = current_employee.first_name if current_employee else request.user.get_full_name() or request.user.username
        note_prefix = f"{date.today().strftime('%m/%d/%Y')} - {completed_by}: Completed task."
        if completion_note:
            note_prefix += f" {completion_note}"

        pending_task.notes = (
            (pending_task.notes + "\n") if pending_task.notes else ""
        ) + note_prefix
        pending_task.is_complete = True
        pending_task.confirmed_is_complete = True
        pending_task.save(update_fields=["certification", "notes", "is_complete", "confirmed_is_complete"])

        if pending_task.certification:
            CertificationNotes.objects.create(
                certification=pending_task.certification,
                date=date.today(),
                user=current_employee or pending_task.employee or Employees.objects.filter(user__is_superuser=True).first() or Employees.objects.first(),
                note=(
                    f"Pending employee task completed for {pending_task.assignee_display}: "
                    f"{pending_task.description}"
                ),
            )

        messages.success(request, "Pending employee task marked complete.")
        return redirect("pending_employee_tasks")

    pending_tasks = list(_pending_employee_task_queryset())
    for task in pending_tasks:
        task.certification_options = _certification_options_for_pending_task(task)

    return render(request, 'pending_employee_tasks.html', {
        "pending_tasks": pending_tasks,
        "pending_tasks_count": len(pending_tasks),
    })


@login_required(login_url='/accounts/login')
def update_pending_employee_task_certification(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid request."}, status=405)

    pending_task = get_object_or_404(
        _pending_employee_task_queryset(),
        id=request.POST.get("pending_task_id"),
    )
    certification_error = _set_pending_task_certification(
        pending_task,
        request.POST.get("certification_id"),
    )
    if certification_error:
        return JsonResponse({"success": False, "message": certification_error}, status=400)

    pending_task.save(update_fields=["certification"])
    if pending_task.certification:
        _apply_certification_display_descriptions([pending_task.certification])
        certification_display = pending_task.certification.display_description
        certification_url = reverse("certifications", args=[pending_task.certification.id])
    else:
        certification_display = ""
        certification_url = ""

    return JsonResponse({
        "success": True,
        "certification_id": pending_task.certification_id,
        "certification_display": certification_display,
        "certification_url": certification_url,
    })


@login_required(login_url='/accounts/login')
def expired_certifications(request):
    certifications = list(_expired_flagged_certification_queryset())
    return render(request, 'expired_certifications.html', {
        "certifications": certifications,
        "certifications_count": len(certifications),
    })


@login_required(login_url='/accounts/login')
@never_cache
def warehouse_home(request):
    send_data = {}
    if Email_Errors.objects.filter(user=request.user.first_name + " " + request.user.last_name).exists():
        send_data['error_message']= Email_Errors.objects.filter(user=request.user.first_name + " " + request.user.last_name).last().error
    Email_Errors.objects.filter(user=request.user.first_name + " " + request.user.last_name).delete()
    pending_pickups = PickupRequest.objects.filter(confirmed=True, is_closed=False).order_by('date')
    if pending_pickups:
        send_data['pending_pickups'] = pending_pickups
    pending_rentals = Rentals.objects.filter(
        requested_off_rent=True
    ).filter(
        Q(off_rent_number__isnull=True) | Q(off_rent_number="")
    )
    if pending_rentals.exists():
        send_data["pending_rentals"] = pending_rentals

    rentals_to_check = []

    for rental in Rentals.objects.filter(off_rent_number__isnull=True, is_closed=False):
        if rental.colorize():
            rentals_to_check.append(rental)

    if rentals_to_check:
        send_data["rentals_to_check"] = rentals_to_check
        send_data["check_rentals"] = len(rentals_to_check)
    employee = Employees.objects.get(user=request.user)
    if request.method == 'POST':
        scheduledtalk_id = request.POST.get('scheduledtalk_id')

        if scheduledtalk_id:
            scheduled_talk = get_object_or_404(
                ScheduledToolboxTalks,
                id=scheduledtalk_id
            )

            has_viewed = ViewedToolboxTalks.objects.filter(
                employee=employee,
                master=scheduled_talk
            ).exists()

            if not has_viewed:
                messages.error(
                    request,
                    "You need to read the toolbox talk first before marking the course complete."
                )
                return redirect('warehouse_home')

            completed_talk, created = CompletedToolboxTalks.objects.get_or_create(
                employee=employee,
                master=scheduled_talk,
                defaults={
                    'date': date.today(),
                    'is_excused': False,
                }
            )
            if not created and completed_talk.is_excused:
                completed_talk.is_excused = False
                completed_talk.date = date.today()
                completed_talk.save(update_fields=['is_excused', 'date'])

            messages.success(request, "Toolbox talk completed.")
            return redirect('warehouse_home')
    if employee.job_title.description == "Painter" or employee.job_title.description == "Superintendent" or employee.job_title.description == "Warehouse":

        toolbox_talks_required = []

        scheduled_qs = ScheduledToolboxTalks.objects.filter(
            date__lte=date.today(),
            date__gte=employee.date_added
        ).filter(
            Q(is_all_employees=True) |
            Q(scheduledtoolboxtalkemployees__employee=employee)
        ).distinct().order_by('date')

        for x in scheduled_qs:

            if CompletedToolboxTalks.objects.filter(employee=employee, master=x, is_excused=False).exists():
                continue

            if x.master:
                talk_description = x.master.description
                talk_display_id = x.master.id
            else:
                talk_description = x.description or "Custom Toolbox Talk"
                talk_display_id = x.id

            english_file = get_uploaded_toolbox_file(x, "English")
            spanish_file = get_uploaded_toolbox_file(x, "Spanish")

            english = english_file['filename'] if english_file else None
            spanish = spanish_file['filename'] if spanish_file else None

            spanish_view = ViewedToolboxTalks.objects.filter(
                employee=employee,
                master=x,
                language="Spanish"
            ).order_by('-date').first()

            english_view = ViewedToolboxTalks.objects.filter(
                employee=employee,
                master=x,
                language="English"
            ).order_by('-date').first()

            toolbox_talks_required.append({
                'id': talk_display_id,
                'item': x.id,
                'description': talk_description,
                'date': x.date,
                'english': english,
                'spanish': spanish,
                'spanish_viewed': bool(spanish_view),
                'spanish_date': spanish_view.date if spanish_view else None,
                'english_viewed': bool(english_view),
                'english_date': english_view.date if english_view else None,
                'can_complete': bool(spanish_view or english_view),
                'notes': x.notes or "",
                'is_all_employees': x.is_all_employees,
            })

        send_data['toolbox_talks_required'] = toolbox_talks_required
        send_data['toolbox_talks_required_count'] = len(toolbox_talks_required)

    return render(request, 'warehouse_home.html', send_data)


@login_required(login_url='/accounts/login')
def admin_home(request):
    send_data = {}
    if request.method == "POST":
        if 'sirius_notes' in request.POST:
            form = SiriusUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                notes = request.POST["sirius_notes"]
                upload_sirius(form, excel_file,notes)
        if 'clockshark_notes' in request.POST:
            form = ClockSharkUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                notes = request.POST["clockshark_notes"]
                # message = upload_clockshark(form, excel_file,notes)
                upload_clockshark(form, excel_file, notes)
                send_data['error_message'] = "Success Uploading ClockShark Hours"
        if 'toolbox_talks' in request.POST:
            form = ToolboxTalksUploadForm(request.POST, request.FILES)
            if form.is_valid():
                excel_file = request.FILES["excel_file"]
                upload_toolbox_talk(form, excel_file, "none")
                send_data['error_message'] = "Success Uploading Toolbox Talks"
        if 'email_test' in request.POST:
            send_data['emailconfirmation'] = True
            Email.sendEmail("Trinity Email Test",
                            "Trinity Test Email sent on " + datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                            ['joe@gerloffpainting.com', 'doug@hrdata.com'], False,"operations@gerloffpainting.com")

    send_data['employees'] = Employees.objects.filter(user__isnull=True, active=True)
    send_data['subs'] = Subcontractors.objects.filter(is_inactive=False)
    send_data['sirius_form'] = SiriusUploadForm()
    send_data['clockshark_form'] = ClockSharkUploadForm()
    send_data['clockshark_form'] = ToolboxTalksUploadForm()
    return render(request, 'admin_home.html', send_data)


@login_required(login_url='/accounts/login')
def import_certifications(request):
    send_data = {
        "reviewed_rows": [],
        "errors": [],
        "warnings": [],
        "import_results": None,
        "mode": "",
    }

    if request.method == "POST":
        uploaded_file = request.FILES.get("certification_file")
        mode = request.POST.get("import_mode")
        send_data["mode"] = mode

        if not uploaded_file:
            messages.error(request, "Please choose a certification import file.")
            return render(request, "import_certifications.html", send_data)

        try:
            rows = _certification_import_rows(uploaded_file)
        except Exception:
            messages.error(request, "The uploaded file could not be read. Please upload an Excel .xlsx file.")
            return render(request, "import_certifications.html", send_data)

        if not rows:
            messages.error(request, "No certification rows were found in the uploaded file.")
            return render(request, "import_certifications.html", send_data)

        reviewed_rows, errors, warnings = _review_certification_import_rows(rows)
        send_data.update({
            "reviewed_rows": reviewed_rows,
            "errors": errors,
            "warnings": warnings,
        })

        if mode == "final":
            if errors:
                messages.error(
                    request,
                    "Final upload was blocked because one or more employees/subcontractors could not be matched.",
                )
                return render(request, "import_certifications.html", send_data)

            import_results = _commit_certification_import(reviewed_rows, request.user)
            send_data["import_results"] = import_results
            send_data["warnings"] = warnings + import_results["warnings"]
            messages.success(request, f"Imported {import_results['certifications_created']} certification(s).")
        else:
            messages.info(request, "Test complete. No certifications were created.")

    return render(request, "import_certifications.html", send_data)


def base(request):
    current_user = request.user
    filteredEmployee = Employees.objects.filter(user=current_user.id).first()
    employee = {}
    if filteredEmployee is not None and filteredEmployee.job_title is not None:
        employee = {'role': filteredEmployee.job_title.description}
    return HttpResponse(json.dumps(employee))



@login_required(login_url='/accounts/login')
def grant_web_access(request,id=None):
    send_data = {}
    employee = None
    if id:
        employee = get_object_or_404(Employees, id=id)
        send_data['selected_employee'] = employee
    send_data['employees'] = Employees.objects.filter(user__isnull=True, pin__isnull=True,active=True).order_by('first_name')
    if request.method == 'POST':
        selected_employee = Employees.objects.get(id=request.POST['select_employee'])
        tester = False
        while tester == False:
            randomPin = random.randint(1000, 9999)
            tester = True
            for x in Employees.objects.filter(user__isnull=True, pin__isnull=False,active=True):
                if x.pin == randomPin:
                    tester = False
                    randomPin = random.randint(1000, 9999)
        selected_employee.pin = randomPin
        selected_employee.save()
        return redirect('admin_home')
    return render(request, 'grant_web_access.html', send_data)


@login_required(login_url='/accounts/login')
def grant_subcontractor_web_access(request):
    send_data = {}
    send_data['subcontractors'] = Subcontractors.objects.filter(is_inactive=False)
    if request.method == 'POST':
        selected_sub = Subcontractors.objects.get(id=request.POST['select_employee'])
        tester = False
        while tester == False:
            randomPin = random.randint(1000, 9999)
            tester = True
            for x in Subcontractors.objects.filter(pin__isnull=False):
                if x.pin == randomPin:
                    tester = False
                    randomPin = random.randint(1000, 9999)
        selected_sub.pin = randomPin
        selected_sub.save()
        return redirect('admin_home')
    return render(request, 'grant_subcontractor_web_access.html', send_data)


# Create your views here.
def register_user(request):
    send_data = {}
    send_data['employees'] = EmployeeLevels.objects.filter(user=None)
    if request.method == 'POST':
        # first_name = request.POST['first_name']
        # last_name = request.POST['last_name']
        username = request.POST['username'].strip()
        password1 = request.POST['password1']
        password2 = request.POST['password2']
        email = request.POST['email']
        if password1 == password2:
            username_exists_in_django_users = User.objects.filter(username__iexact=username).exists() if username else False
            username_exists_in_subcontractors = Subcontractors.objects.filter(username__iexact=username).exists() if username else False
            username_exists_in_subcontractor_employees = Subcontractor_Employees.objects.filter(username__iexact=username).exists() if username else False

            if not username or username_exists_in_django_users or username_exists_in_subcontractors or username_exists_in_subcontractor_employees:
                messages.error(request, 'USERNAME ALREADY IN USE. Please choose a different username.')
                return redirect('register_user')
            else:
                user = User.objects.create_user(username=username, password=password1, email=email,
                                                first_name=Employees.objects.get(
                                                    id=request.POST['select_employee']).first_name,
                                                last_name=Employees.objects.get(
                                                    id=request.POST['select_employee']).last_name, is_active=False)
                user.save();
                employee = Employees.objects.get(id=request.POST['select_employee'])
                employee.user = user
                employee.save()
                return redirect('login')
        else:
            messages.info(request, 'password not matching...')
            return redirect('register_user')

    else:
        return render(request, 'register.html', send_data)


def import_csv2(request):
    equipment.models.InventoryItems4.objects.all().delete()
    equipment.models.InventoryItems3.objects.all().delete()
    equipment.models.InventoryItems2.objects.all().delete()
    equipment.models.InventoryItems.objects.all().delete()
    equipment.models.InventoryType.objects.all().delete()
    employees.models.MetricLevels.objects.all().delete()
    employees.models.MetricCategories.objects.all().delete()
    employees.models.TrainingTopic.objects.all().delete()
    employees.models.Metrics.objects.all().delete()

    with open("c:/sql_backup/certificationactionrequired.csv") as f:
        current_table = employees.models.CertificationActionRequired
        current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(2):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "action":
                        b = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 2:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], action=row[b])

    with open("c:/sql_backup/certificationcategories.csv") as f:
        current_table = employees.models.CertificationCategories
        current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(2):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "description":
                        b = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 2:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], description=row[b])

    with open("c:/sql_backup/employeelevels.csv") as f:
        current_table = employees.models.EmployeeLevels
        current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(3):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "description":
                        b = x
                        found = found + 1
                    if row[x] == "pay_rate":
                        c = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 3:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], description=row[b], pay_rate=row[c])

    with open("c:/sql_backup/employeetitles.csv") as f:
        current_table = employees.models.EmployeeTitles
        current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(2):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "description":
                        b = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 2:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], description=row[b])

    with open("c:/sql_backup/exam.csv") as f:
        current_table = employees.models.Exam
        current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(4):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "description":
                        b = x
                        found = found + 1
                    if row[x] == "details":
                        c = x
                        found = found + 1
                    if row[x] == "max_score":
                        d = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 4:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], description=row[b], details=row[c], max_score=row[d])

    with open("c:/sql_backup/inventorytype.csv") as f:
        current_table = equipment.models.InventoryType
        # current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(3):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "type":
                        b = x
                        found = found + 1
                    if row[x] == "is_active":
                        c = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 3:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], type=row[b], is_active=row[c])

    with open("c:/sql_backup/inventoryitems.csv") as f:
        current_table = equipment.models.InventoryItems
        # current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(3):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "name":
                        b = x
                        found = found + 1
                    if row[x] == "type_id":
                        c = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 3:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], name=row[b],
                                             type=equipment.models.InventoryType.objects.get(id=row[c]))

    with open("c:/sql_backup/inventoryitems2.csv") as f:
        current_table = equipment.models.InventoryItems2
        # current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(3):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "name":
                        b = x
                        found = found + 1
                    if row[x] == "type_id":
                        c = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 3:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], name=row[b],
                                             type=equipment.models.InventoryItems.objects.get(id=row[c]))

    with open("c:/sql_backup/inventoryitems3.csv") as f:
        current_table = equipment.models.InventoryItems3
        # current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(3):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "name":
                        b = x
                        found = found + 1
                    if row[x] == "type_id":
                        c = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 3:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], name=row[b],
                                             type=equipment.models.InventoryItems2.objects.get(id=row[c]))

    with open("c:/sql_backup/inventoryitems4.csv") as f:
        current_table = equipment.models.InventoryItems4
        # current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(3):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "name":
                        b = x
                        found = found + 1
                    if row[x] == "type_id":
                        c = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 3:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], name=row[b],
                                             type=equipment.models.InventoryItems3.objects.get(id=row[c]))

    with open("c:/sql_backup/jobnumbers.csv") as f:
        current_table = jobs.models.JobNumbers
        current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(3):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "letter":
                        b = x
                        found = found + 1
                    if row[x] == "number":
                        c = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 3:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(letter=row[b], number=row[c])

    with open("c:/sql_backup/metrics.csv") as f:
        current_table = employees.models.Metrics
        # current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(2):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "description":
                        b = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 2:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], description=row[b])

    with open("c:/sql_backup/metriclevels.csv") as f:
        current_table = employees.models.MetricLevels
        # current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(3):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "level_id":
                        b = x
                        found = found + 1
                    if row[x] == "metric_id":
                        c = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 3:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], level=employees.models.EmployeeLevels.objects.get(id=row[b]),
                                             metric=employees.models.Metrics.objects.get(id=row[c]))

    with open("c:/sql_backup/metriccategories.csv") as f:
        current_table = employees.models.MetricCategories
        # current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(4):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "number":
                        b = x
                        found = found + 1
                    if row[x] == "description":
                        c = x
                        found = found + 1
                    if row[x] == "metric_id":
                        d = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 4:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], number=row[b], description=row[c],
                                             metric=employees.models.Metrics.objects.get(id=row[d]))

    with open("c:/sql_backup/productioncategory.csv") as f:
        current_table = employees.models.ProductionCategory
        current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(8):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "item1":
                        b = x
                        found = found + 1
                    if row[x] == "item2":
                        c = x
                        found = found + 1
                    if row[x] == "item3":
                        d = x
                        found = found + 1
                    if row[x] == "task":
                        e = x
                        found = found + 1
                    if row[x] == "unit1":
                        f = x
                        found = found + 1
                    if row[x] == "unit2":
                        g = x
                        found = found + 1
                    if row[x] == "unit3":
                        h = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 8:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], item1=row[b], item2=row[c], item3=row[d], task=row[e],
                                             unit1=row[f], unit2=row[g], unit3=row[h])

    with open("c:/sql_backup/tmpricesmaster.csv") as f:
        current_table = changeorder.models.TMPricesMaster
        current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(5):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "category":
                        b = x
                        found = found + 1
                    if row[x] == "item":
                        c = x
                        found = found + 1
                    if row[x] == "unit":
                        d = x
                        found = found + 1
                    if row[x] == "rate":
                        e = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 5:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], category=row[b], item=row[c], unit=row[d], rate=row[e])

    with open("c:/sql_backup/trainingtopic.csv") as f:
        current_table = employees.models.TrainingTopic
        # current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(5):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "description":
                        b = x
                        found = found + 1
                    if row[x] == "details":
                        c = x
                        found = found + 1
                    if row[x] == "assessment_category_id":
                        d = x
                        found = found + 1
                    if row[x] == "assessment_category1_id":
                        e = x
                        found = found + 1

                line_count1 = line_count1 + 1
                if found != 5:
                    raise ValueError('A very specific bad thing happened.')
            else:
                new_item = current_table.objects.create(id=row[a], description=row[b], details=row[c])
                if row[d] != '':
                    new_item.assessment_category = employees.models.Metrics.objects.get(id=row[d])
                if row[e] != '':
                    new_item.assessment_category1 = employees.models.Metrics.objects.get(id=row[e])
    with open("c:/sql_backup/vendorcategory.csv") as f:
        current_table = equipment.models.VendorCategory
        current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(2):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "category":
                        b = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 2:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], category=row[b])

    with open("c:/sql_backup/writeupdefaults.csv") as f:
        current_table = employees.models.WriteUpDefaults
        current_table.objects.all().delete()
        reader = csv.reader(f)
        line_count1 = 0
        found = 0
        for row in reader:
            if line_count1 == 0:
                for x in range(2):
                    if row[x] == "id":
                        a = x
                        found = found + 1
                    if row[x] == "description":
                        b = x
                        found = found + 1
                line_count1 = line_count1 + 1
                if found != 2:
                    raise ValueError('A very specific bad thing happened.')
            else:
                current_table.objects.create(id=row[a], description=row[b])


    return render(request, 'index.html')


def reset_databases(request):
    # git hub2
    if request.user.first_name == "Joe" and request.user.last_name == "Gerloff":
        TMList.objects.all().delete()
        TMProposal.objects.all().delete()
        EWTicket.objects.all().delete()
        EWT.objects.all().delete()
        TempRecipients.objects.all().delete()
        ChangeOrderNotes.objects.all().delete()
        ChangeOrders.objects.all().delete()
        Signature.objects.all().delete()
        InventoryNotes.objects.all().delete()
        Inventory.objects.all().delete()
        OutgoingItem.objects.all().delete()
        OutgoingWallcovering.objects.all().delete()
        Packages.objects.all().delete()
        ReceivedItems.objects.all().delete()
        WallcoveringDelivery.objects.all().delete()
        OrderItems.objects.all().delete()
        WallcoveringPricing.objects.all().delete()
        Wallcovering.objects.all().delete()
        Orders.objects.all().delete()
        RentalNotes.objects.all().delete()
        Rentals.objects.all().delete()
        SubmittalNotes.objects.all().delete()
        SubmittalItems.objects.all().delete()
        Submittals.objects.all().delete()
        ClientEmployees.objects.all().delete()  # dangerous
        Clients.objects.all().delete()  # dangerous
        JobNotes.objects.all().delete()
        Jobs.objects.all().delete()
    return redirect("/")


def create_folders(request):
    for x in Inventory.objects.all():
        createfolder("equipment/" + str(x.id))

    return render(request, 'index.html')


def customize(request):
    for x in ChangeOrderNotes.objects.all():
        if x.user == "Bridgette Clause":
            x.user = 12
        elif x.user == "Charity Archibald":
            x.user = 14
        elif x.user == "Steve Beaudoin":
            x.user = 21
        elif x.user == "Edward Diggs":
            x.user = 26
        elif x.user == "Joe Gerloff":
            x.user = 3
        elif x.user == "Anthony Taroli":
            x.user = 5
        elif x.user == "D'Angelo Smith":
            x.user = 40
        else:
            print("nothing")
        x.save()

    for x in RentalNotes.objects.all():
        if x.user == "Bridgette Clause":
            x.user = 12
        elif x.user == "Charity Archibald":
            x.user = 14
        elif x.user == "Steve Beaudoin":
            x.user = 21
        elif x.user == "Edward Diggs":
            x.user = 26
        elif x.user == "Joe Gerloff":
            x.user = 3
        elif x.user == "Anthony Taroli":
            x.user = 5
        elif x.user == "D'Angelo Smith":
            x.user = 40
        else:
            print("nothing")
        x.save()

    return redirect('/')


def import_csv(request):
    with open("c:/sql_backup/workorderimport.csv", encoding='utf-8-sig') as f:
        current_table = employees.models.CertificationActionRequired
        current_table.objects.all().delete()
        reader = csv.reader(f)
        for row in reader:
            try:
                job = Jobs.objects.get(job_number=row[0])
                if row[1]:
                    job.is_work_order_done = True
                job.save()
            except:
                print(row[0])
        return render(request, 'index.html')

def upload_sirius(form, excel_file,notes):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    created = 0
    skipped = 0
    with transaction.atomic():
        for row in sheet.iter_rows(min_row=1, values_only=True):
            (
                job_number,
                hours,
            ) = row

            # REQUIRED FIELDS CHECK
            # if not first_name or not last_name or not employer:
            #     skipped += 1
            #     continue

            # Resolve Foreign Keys safely
            # level = None
            # if level_name:
            #     level = EmployeeLevels.objects.filter(name=level_name).first()

            # job_title = None
            # if job_title_name:
            #     job_title = EmployeeTitles.objects.filter(description=job_title_name).first()
            job = Jobs.objects.filter(job_number=job_number).first()
            SiriusHours.objects.create(
                job=job,
                job_number=job_number,
                date=date.today(),
                hours=hours,
                notes = notes
            )

            created += 1

#old as of 6.2.26, i changed it to .csv functionality
def upload_clockshark_old(form, excel_file,notes):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    created = 0
    skipped = 0
    a=0
    with transaction.atomic():
        for row in sheet.iter_rows(min_row=2, values_only=True):#24
            (
                first_name,
                last_name,
                ignore,
                ignore,
                job_name,
                job_number,
                ignore,
                ignore,
                start_raw,
                end_raw,
                lunch,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
            ) = row
            if a==0:
                oldest_date = start_raw
                newest_date = start_raw
            else:
                newest_date = start_raw
            a += 1

        if oldest_date and timezone.is_naive(oldest_date):
            oldest_date = timezone.make_aware(oldest_date, timezone.get_current_timezone())
        message = []
        if newest_date and timezone.is_naive(newest_date):
            newest_date = timezone.make_aware(newest_date, timezone.get_current_timezone())
        oldest_day = oldest_date.date()
        newest_day = newest_date.date()
        ClockSharkTimeEntry.objects.filter(work_day__gte=oldest_date, work_day__lte=newest_date).delete()
        with transaction.atomic():
            for row in sheet.iter_rows(min_row=2, values_only=True):  # 24
                (
                    employee_first_name,
                    employee_last_name,
                    ignore,
                    ignore,
                    job_name,
                    job_number,
                    ignore,
                    ignore,
                    start_raw,
                    end_raw,
                    lunch,
                    ignore,
                    ignore,
                    ignore,
                    ignore,
                    ignore,
                    ignore,
                    ignore,
                    minutes,
                    ignore,
                    ignore,
                    ignore,
                    ignore,
                    ignore,
                ) = row
                a=a+1
                if job_number:
                    job_number=str(job_number).strip()
                    job_number=job_number[:5]
                if job_name:
                    job_name=str(job_name).strip()
                if job_name.endswith("*"):
                    job_name=job_name[:-1]
                if lunch in (None, "", 0):
                    lunch = Decimal("0")
                else:
                    lunch = Decimal(str(lunch))
                clock_in_time=start_raw
                clock_out_time=end_raw
                job = Jobs.objects.filter(job_number=job_number).first()
                if clock_in_time:
                    work_day = clock_in_time.date()
                if clock_out_time:
                    work_day = clock_out_time.date()
                clockshark_id = f"{employee_first_name}|{employee_last_name}|{job_name}|{work_day}"
                if clock_in_time and timezone.is_naive(clock_in_time):
                    clock_in_time = timezone.make_aware(clock_in_time, timezone.get_current_timezone())
                if clock_out_time and timezone.is_naive(clock_out_time):
                    clock_out_time = timezone.make_aware(clock_out_time, timezone.get_current_timezone())

                if job:
                    if not job.is_active:
                        if not JobNotes.objects.filter(job_number=job,
                                                       note__contains="Changed Status to Active").exists():
                            job.is_active = True
                            job.save()
                            JobNotes.objects.create(job_number=job,
                                                    note="Changed Status to Active From Clock Shark Import",
                                                    type="auto_start_date_note",
                                                    user=Employees.objects.filter().first(), date=date.today())

                if work_day < date.today():
                    ClockSharkTimeEntry.objects.create(clockshark_id=clockshark_id, job_name=job_name,
                                                       employee_first_name=employee_first_name,
                                                       employee_last_name=employee_last_name, work_day=work_day,
                                                       clock_in=clock_in_time, job=job, clock_out=clock_out_time,
                                                       hours=minutes / 60, hours_adjust_note="AUTO IMPORT",lunch=lunch)
                else:
                    if clock_in_time and clock_out_time:
                        ClockSharkTimeEntry.objects.create(lunch=lunch,clockshark_id=clockshark_id, job_name=job_name,
                                                           employee_first_name=employee_first_name,
                                                           employee_last_name=employee_last_name, work_day=work_day,
                                                           clock_in=clock_in_time,job=job, clock_out=clock_out_time, hours=minutes/60, hours_adjust_note="AUTO IMPORT")

    # return message

def parse_clockshark_datetime(value):
    """
    Handles Excel datetime values and common CSV datetime strings.
    Adjust formats if ClockShark exports a different date format.
    """
    if not value:
        return None

    if isinstance(value, datetime.datetime):
        return value

    value = str(value).strip()

    formats = [
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %I:%M:%S %p",
        "%m/%d/%Y %H:%M",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
    ]

    for fmt in formats:
        try:
            return datetime.datetime.strptime(value, fmt)
        except ValueError:
            pass

    raise ValueError(f"Could not parse datetime: {value}")


def get_clockshark_rows(uploaded_file):
    """
    Returns rows from either CSV or Excel as a list of tuples/lists.
    Skips the header row.
    """
    filename = uploaded_file.name.lower()

    if filename.endswith(".csv"):
        decoded_file = uploaded_file.read().decode("utf-8-sig").splitlines()
        reader = csv.reader(decoded_file)
        rows = list(reader)
        return rows[1:]  # skip header

    wb = openpyxl.load_workbook(uploaded_file, data_only=True)
    sheet = wb.active
    return list(sheet.iter_rows(min_row=2, values_only=True))

def upload_clockshark(form, excel_file, notes):

    rows = get_clockshark_rows(excel_file)

    created = 0
    skipped = 0
    a = 0
    oldest_date = None
    newest_date = None

    with transaction.atomic():
        for row in rows:
            (
                first_name,
                last_name,
                ignore,
                ignore,
                ignore,
                job_name,
                job_number,
                ignore,
                ignore,
                start_raw,
                end_raw,
                lunch,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
                ignore,
            ) = row

            start_raw = parse_clockshark_datetime(start_raw)

            if not start_raw:
                continue

            if a == 0:
                oldest_date = start_raw
                newest_date = start_raw
            else:
                newest_date = start_raw

            a += 1

        if oldest_date and timezone.is_naive(oldest_date):
            oldest_date = timezone.make_aware(oldest_date, timezone.get_current_timezone())
        message = []
        if newest_date and timezone.is_naive(newest_date):
            newest_date = timezone.make_aware(newest_date, timezone.get_current_timezone())

        if not oldest_date or not newest_date:
            print("No valid ClockShark dates found.")
            return

        oldest_day = oldest_date.date()
        newest_day = newest_date.date()

        ClockSharkTimeEntry.objects.filter(
            work_day__gte=oldest_day,
            work_day__lte=newest_day
        ).delete()
        with transaction.atomic():
            for row in rows:  # 24
                (
                    employee_first_name,
                    employee_last_name,
                    ignore,
                    ignore,
                    ignore,
                    job_name,
                    job_number,
                    ignore,
                    ignore,
                    start_raw,
                    end_raw,
                    lunch,
                    ignore,
                    ignore,
                    ignore,
                    ignore,
                    ignore,
                    ignore,
                    ignore,
                    minutes,
                    ignore,
                    ignore,
                    ignore,
                    ignore,
                    ignore,
                ) = row
                a=a+1
                if job_number:
                    job_number=str(job_number).strip()
                    job_number=job_number[:5]
                if job_name:
                    job_name = str(job_name).strip()
                    if job_name.endswith("*"):
                        job_name = job_name[:-1]
                else:
                    job_name = ""
                if lunch in (None, "", 0):
                    lunch = Decimal("0")
                else:
                    lunch = Decimal(str(lunch))
                if minutes in (None, ""):
                    minutes = Decimal("0")
                else:
                    minutes = Decimal(str(minutes))
                clock_in_time = parse_clockshark_datetime(start_raw)
                clock_out_time = parse_clockshark_datetime(end_raw)
                job = Jobs.objects.filter(job_number=job_number).first()
                work_day = None

                if clock_in_time:
                    work_day = clock_in_time.date()
                elif clock_out_time:
                    work_day = clock_out_time.date()

                if not work_day:
                    print("skipped row with no work day: " + str(a))
                    continue

                clockshark_id = f"{employee_first_name}|{employee_last_name}|{job_name}|{work_day}"
                if clock_in_time and timezone.is_naive(clock_in_time):
                    clock_in_time = timezone.make_aware(clock_in_time, timezone.get_current_timezone())
                if clock_out_time and timezone.is_naive(clock_out_time):
                    clock_out_time = timezone.make_aware(clock_out_time, timezone.get_current_timezone())

                if job:
                    if not job.is_active:
                        if not JobNotes.objects.filter(job_number=job,
                                                       note__contains="Changed Status to Active").exists():
                            job.is_active = True
                            job.save()
                            JobNotes.objects.create(job_number=job,
                                                    note="Changed Status to Active From Clock Shark Import",
                                                    type="auto_start_date_note",
                                                    user=Employees.objects.filter().first(), date=date.today())
                else:
                    # message.append({'message': "couldn't find " + job_name + " row: " + str(a)})
                    print("couldn't find " + job_name + " row: " + str(a))
                if work_day < date.today():
                    ClockSharkTimeEntry.objects.create(clockshark_id=clockshark_id, job_name=job_name,
                                                       employee_first_name=employee_first_name,
                                                       employee_last_name=employee_last_name, work_day=work_day,
                                                       clock_in=clock_in_time, job=job, clock_out=clock_out_time,
                                                       hours=minutes / 60, hours_adjust_note="AUTO IMPORT",lunch=lunch)
                else:
                    if clock_in_time and clock_out_time:
                        ClockSharkTimeEntry.objects.create(lunch=lunch,clockshark_id=clockshark_id, job_name=job_name,
                                                           employee_first_name=employee_first_name,
                                                           employee_last_name=employee_last_name, work_day=work_day,
                                                           clock_in=clock_in_time,job=job, clock_out=clock_out_time, hours=minutes/60, hours_adjust_note="AUTO IMPORT")
                    else:
                        # message.append({'message': "skipped " + clockshark_id + " row" + str(a)})
                        print("skipped " + clockshark_id + " row" + str(a))
    # return message


def upload_toolbox_talk(form, excel_file,notes):
    wb = openpyxl.load_workbook(excel_file)
    sheet = wb.active
    created = 0
    skipped = 0
    with transaction.atomic():
        for row in sheet.iter_rows(min_row=1, values_only=True):
            (
                description,
            ) = row

            newitem = ToolboxTalks.objects.create(
                description=description,
            )
            createfolder("toolbox_talks/" + str(newitem.id))
            createfolder("toolbox_talks/" + str(newitem.id) + "/English")
            createfolder("toolbox_talks/" + str(newitem.id) + "/Spanish")


def tm_prices_master(request):

    if request.method == "POST":

        action = request.POST.get("action")

        if action == "update":

            item_id = request.POST.get("id")
            item = TMPricesMaster.objects.get(id=item_id)

            item.item = request.POST.get("item")
            item.unit = request.POST.get("unit")
            item.rate = request.POST.get("rate")

            item.save()

        elif action == "delete":

            item_id = request.POST.get("id")
            TMPricesMaster.objects.filter(id=item_id).delete()

        elif action == "add":

            TMPricesMaster.objects.create(
                category=request.POST.get("category"),
                item=request.POST.get("item"),
                unit=request.POST.get("unit"),
                rate=request.POST.get("rate")
            )

        return redirect("tm_prices_master")

    categories = [
        "Labor",
        "Material",
        "Equipment",
        "Inventory",
        "Misc",
        "Bond",
        "Sundries"
    ]

    data = {}

    for c in categories:
        data[c] = TMPricesMaster.objects.filter(category=c).order_by("item")

    return render(request, "tm_prices_master.html", {"data": data})


def parse_decimal(value):
    if value in [None, ""]:
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, AttributeError, TypeError, ValueError):
        return None


def parse_date(value):
    if value in [None, ""]:
        return None

    if isinstance(value, datetime.datetime):
        return value.date()

    if isinstance(value, datetime.date):
        return value

    for fmt in ["%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"]:
        try:
            return datetime.datetime.strptime(str(value).strip(), fmt).date()
        except ValueError:
            continue

    return None


def parse_bool(value):
    if isinstance(value, bool):
        return value

    if value in [None, ""]:
        return False

    return str(value).strip().lower() in ["true", "1", "yes", "y"]


@transaction.atomic
def import_change_orders(request):
    if request.method == "POST":
        if "file" not in request.FILES:
            messages.error(request, "No file was uploaded.")
            return redirect("admin_home")

        excel_file = request.FILES["file"]

        try:
            wb = load_workbook(excel_file, data_only=True)
            ws = wb.active

            created_count = 0
            updated_count = 0
            skipped_rows = []

            headers = [cell.value for cell in ws[1]]

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_data = dict(zip(headers, row))

                excel_job_number = row_data.get("job_number")
                cop_number = row_data.get("cop_number")

                if not excel_job_number:
                    skipped_rows.append(f"Row {row_num}: missing job_number")
                    continue

                if cop_number in [None, ""]:
                    skipped_rows.append(f"Row {row_num}: missing cop_number")
                    continue

                job = Jobs.objects.filter(job_number=str(excel_job_number).strip()).first()
                if not job:
                    skipped_rows.append(f"Row {row_num}: job not found ({excel_job_number})")
                    continue

                defaults = {
                    "description": row_data.get("description") or "",
                    "price": parse_decimal(row_data.get("price")),
                    "date_sent": parse_date(row_data.get("date_sent")),
                    "date_approved": parse_date(row_data.get("date_approved")),
                    "gc_number": str(row_data.get("gc_number")).strip() if row_data.get("gc_number") not in [None, ""] else "",
                    "is_closed": parse_bool(row_data.get("is_closed")),
                    "notes": "Imported from Management Console",
                    "is_work_complete": parse_bool(row_data.get("is_work_complete")),
                    "is_t_and_m": parse_bool(row_data.get("is_t_and_m")),
                    "is_ticket_signed": parse_bool(row_data.get("is_ticket_signed")),
                    "date_signed": parse_date(row_data.get("date_signed")),
                    "date_week_ending": parse_date(row_data.get("date_week_ending")),
                    "is_approved_to_bill": parse_bool(row_data.get("is_approved_to_bill")),
                    "originated_in_management_console": parse_bool(row_data.get("originated_in_management_console")),
                    "is_approved": parse_bool(row_data.get("is_approved")),
                }

                obj, created = ChangeOrders.objects.update_or_create(
                    job_number=job,
                    cop_number=int(cop_number),
                    defaults=defaults
                )
                createfolder(
                    "changeorder/" + str(obj.job_number.job_number) + " COP #" + str(obj.cop_number))

                excel_note = row_data.get("Notes")

                if excel_note:
                    ChangeOrderNotes.objects.create(
                        cop_number=obj,  # THIS is the ChangeOrder instance
                        date=date.today(),
                        user=Employees.objects.get(user=request.user),
                        note="Imported from Management Console"
                    )
                    ChangeOrderNotes.objects.create(
                        cop_number=obj,  # THIS is the ChangeOrder instance
                        date=date.today(),
                        user=Employees.objects.get(user=request.user),
                        note=str(excel_note)
                    )
                if created:
                    created_count += 1
                else:
                    updated_count += 1

            messages.success(
                request,
                f"Import complete. Created: {created_count}, Updated: {updated_count}, Skipped: {len(skipped_rows)}"
            )

            for msg in skipped_rows[:20]:
                messages.warning(request, msg)

        except Exception as e:
            messages.error(request, f"Import failed: {str(e)}")

        return redirect("admin_home")

    return render(request, "import_change_orders.html")



def job_prices(request, job_number):
    job = get_object_or_404(Jobs, job_number=job_number)

    if request.method == "POST":
        if "delete_job_charge" in request.POST:
            charge_id = request.POST.get("job_charge_id")

            charge = JobCharges2.objects.filter(
                id=charge_id,
                job=job
            ).first()

            if charge:
                deleted_item = charge.item
                charge.delete()
                messages.success(request, f"Deleted job charge: {deleted_item}.")
            else:
                messages.error(request, "Could not find that job charge.")

            return redirect("job_prices", job_number=job.job_number)
        # -----------------------------------
        # Save job-specific TMPricesMaster override
        # -----------------------------------
        if "master_id" in request.POST:
            master_id = request.POST.get("master_id")
            raw_rate = request.POST.get("rate", "").replace(",", "").strip()

            try:
                new_rate = Decimal(raw_rate).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP
                )
            except (InvalidOperation, TypeError):
                messages.error(request, "Please enter a valid rate.")
                return redirect("job_prices", job_number=job.job_number)

            master = get_object_or_404(TMPricesMaster, id=master_id)

            existing_job_price = JobPrices.objects.filter(
                master=master,
                job_number=job
            ).first()

            current_rate = existing_job_price.rate if existing_job_price else master.rate

            if current_rate != new_rate:
                if existing_job_price:
                    existing_job_price.rate = new_rate
                    existing_job_price.save()
                else:
                    JobPrices.objects.create(
                        master=master,
                        rate=new_rate,
                        job_number=job
                    )
                messages.success(request, f"Saved override for {master.item}.")
            else:
                messages.info(request, f"No change made for {master.item}.")

            return redirect("job_prices", job_number=job.job_number)

        # -----------------------------------
        # Add new JobCharges2 row
        # -----------------------------------
        if "add_job_charge" in request.POST:
            category = "Misc"
            item = request.POST.get("charge_item", "").strip()
            unit = request.POST.get("charge_unit", "").strip()
            raw_rate = request.POST.get("charge_rate", "").replace(",", "").strip()

            if not category or not item or not unit or raw_rate == "":
                messages.error(request, "Please fill out all new charge fields.")
                return redirect("job_prices", job_number=job.job_number)

            try:
                rate = Decimal(raw_rate).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP
                )
            except (InvalidOperation, TypeError):
                messages.error(request, "Please enter a valid charge rate.")
                return redirect("job_prices", job_number=job.job_number)

            JobCharges2.objects.create(
                job=job,
                category=category,
                item=item,
                unit=unit,
                rate=rate
            )

            messages.success(request, f"Added job charge: {item}.")
            return redirect("job_prices", job_number=job.job_number)

    rows = []
    for master in TMPricesMaster.objects.all().order_by("category", "item"):
        override = JobPrices.objects.filter(
            master=master,
            job_number=job
        ).first()

        rows.append({
            "master_id": master.id,
            "category": master.category,
            "item": master.item,
            "unit": master.unit,
            "master_rate": master.rate,
            "current_rate": override.rate if override else master.rate,
            "has_override": bool(override),
        })

    job_charges = JobCharges2.objects.filter(job=job).order_by("category", "item")

    return render(
        request,
        "job_prices.html",
        {
            "job": job,
            "rows": rows,
            "job_charges": job_charges,
        }
    )




def export_jobs_ar_closed_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="jobs_ar_closed.csv"'

    writer = csv.writer(response)

    # Header
    writer.writerow(['Job Number', 'AR Closed Date'])

    # Only jobs with AR Closed Date, sorted by date
    jobs = Jobs.objects.filter(
        ar_closed_date__isnull=False
    ).order_by('ar_closed_date')

    for job in jobs:
        writer.writerow([
            job.job_number,
            job.ar_closed_date.strftime('%m/%d/%Y')  # Excel-friendly format
        ])

    return response


def reassign_vendor_foreign_keys(old_vendor, new_vendor):
    """
    Move all known Vendors foreign keys from old_vendor to new_vendor.
    Also moves all VendorContacts from old_vendor to new_vendor.
    """

    Rentals.objects.filter(company=old_vendor).update(company=new_vendor)

    Wallcovering.objects.filter(vendor=old_vendor).update(vendor=new_vendor)

    Inventory.objects.filter(purchased_from=old_vendor).update(purchased_from=new_vendor)

    Inventory.objects.filter(service_vendor=old_vendor).update(service_vendor=new_vendor)

    Orders.objects.filter(vendor=old_vendor).update(vendor=new_vendor)

    # Important: move the vendor contacts too
    VendorContact.objects.filter(company=old_vendor).update(company=new_vendor)


def reassign_vendor_contact_foreign_keys(old_contact, new_contact):
    """
    Move all known VendorContact foreign keys from old_contact to new_contact.
    """

    Rentals.objects.filter(rep=old_contact).update(rep=new_contact)


def clear_vendor_contact_foreign_keys(contact):
    """
    Clear VendorContact foreign keys before deleting a contact.
    Right now, only Rentals.rep points to VendorContact and it is nullable.
    """

    Rentals.objects.filter(rep=contact).update(rep=None)

def get_vendor_reference_counts(vendor):
    return {
        "rentals": Rentals.objects.filter(company=vendor).count(),
        "wallcoverings": Wallcovering.objects.filter(vendor=vendor).count(),
        "inventory_purchased_from": Inventory.objects.filter(purchased_from=vendor).count(),
        "inventory_service_vendor": Inventory.objects.filter(service_vendor=vendor).count(),
        "orders": Orders.objects.filter(vendor=vendor).count(),
        "vendor_contacts": VendorContact.objects.filter(company=vendor).count(),
    }


def get_vendor_contact_reference_counts(contact):
    return {
        "rentals": Rentals.objects.filter(rep=contact).count(),
    }

def vendor_management(request):
    send_data = {}

    categories = VendorCategory.objects.all().order_by('category')
    send_data['categories'] = categories

    selected_category_id = request.GET.get('category_id') or request.POST.get('category_id')
    selected_vendor_id = request.GET.get('vendor_id') or request.POST.get('vendor_id')

    selected_category = None
    selected_vendor = None
    vendors = Vendors.objects.none()
    vendor_contacts = VendorContact.objects.none()

    if selected_category_id:
        selected_category = VendorCategory.objects.get(id=selected_category_id)
        vendors = Vendors.objects.filter(category=selected_category).order_by('company_name')

    if selected_vendor_id:
        selected_vendor = Vendors.objects.get(id=selected_vendor_id)
        vendor_contacts = VendorContact.objects.filter(company=selected_vendor).order_by('name')

    send_data['selected_category'] = selected_category
    send_data['selected_category_id'] = selected_category_id
    send_data['vendors'] = vendors
    send_data['selected_vendor'] = selected_vendor
    send_data['selected_vendor_id'] = selected_vendor_id
    send_data['vendor_contacts'] = vendor_contacts

    if request.method == "POST":
        action = request.POST.get('action')

        # ------------------------------------------------------------
        # OPTION 1: MERGE VENDORS
        # ------------------------------------------------------------
        if action == "merge_vendors":
            old_vendor_id = request.POST.get('old_vendor_id')
            new_vendor_id = request.POST.get('new_vendor_id')

            if not old_vendor_id or not new_vendor_id:
                messages.error(request, "Please select both vendors.")
                return redirect(f"{request.path}?category_id={selected_category_id}")

            if old_vendor_id == new_vendor_id:
                messages.error(request, "You cannot merge a vendor into itself.")
                return redirect(f"{request.path}?category_id={selected_category_id}")

            old_vendor = Vendors.objects.get(id=old_vendor_id)
            new_vendor = Vendors.objects.get(id=new_vendor_id)

            conflicts = vendor_conflicts(old_vendor, new_vendor)

            # If user has not confirmed yet, show conflicts first.
            if request.POST.get('confirm_merge') != "YES" and conflicts:
                send_data['old_vendor_reference_counts'] = get_vendor_reference_counts(old_vendor)
                send_data['vendor_merge_conflicts'] = conflicts
                send_data['merge_old_vendor'] = old_vendor
                send_data['merge_new_vendor'] = new_vendor
                return render(request, 'vendor_management.html', send_data)

            old_vendor_name = old_vendor.company_name
            new_vendor_name = new_vendor.company_name
            with transaction.atomic():
                reassign_vendor_foreign_keys(old_vendor, new_vendor)
                old_vendor.delete()

            messages.success(
                request,
                f"{old_vendor_name} was merged into {new_vendor_name}."
            )

            return redirect(f"{request.path}?category_id={selected_category_id}")

        # ------------------------------------------------------------
        # OPTION 2: UPDATE VENDOR
        # ------------------------------------------------------------
        elif action == "update_vendor":
            vendor_id = request.POST.get('vendor_id')

            vendor = Vendors.objects.get(id=vendor_id)
            vendor.company_name = request.POST.get('company_name', '').strip()
            vendor.company_phone = request.POST.get('company_phone', '').strip()
            vendor.company_email = request.POST.get('company_email', '').strip()

            new_category_id = request.POST.get('vendor_category')
            if new_category_id:
                vendor.category = VendorCategory.objects.get(id=new_category_id)

            vendor.save()

            messages.success(request, "Vendor information was updated.")

            return redirect(
                f"{request.path}?category_id={vendor.category.id}&vendor_id={vendor.id}"
            )

        # ------------------------------------------------------------
        # ADD / UPDATE VENDOR CONTACT
        # ------------------------------------------------------------
        elif action == "save_vendor_contact":
            vendor_id = request.POST.get('vendor_id')
            contact_id = request.POST.get('contact_id')

            vendor = Vendors.objects.get(id=vendor_id)

            contact_name = request.POST.get('contact_name', '').strip()
            contact_phone = request.POST.get('contact_phone', '').strip()
            contact_email = request.POST.get('contact_email', '').strip()

            if not contact_name:
                messages.error(request, "Contact name is required.")
                return redirect(
                    f"{request.path}?category_id={vendor.category.id}&vendor_id={vendor.id}"
                )

            if contact_id == "add_new":
                VendorContact.objects.create(
                    company=vendor,
                    name=contact_name,
                    phone=contact_phone,
                    email=contact_email
                )
                messages.success(request, "Vendor contact was added.")
            else:
                contact = VendorContact.objects.get(id=contact_id, company=vendor)
                contact.name = contact_name
                contact.phone = contact_phone
                contact.email = contact_email
                contact.save()
                messages.success(request, "Vendor contact was updated.")

            return redirect(
                f"{request.path}?category_id={vendor.category.id}&vendor_id={vendor.id}"
            )

        # ------------------------------------------------------------
        # DELETE VENDOR CONTACT
        # ------------------------------------------------------------
        elif action == "delete_vendor_contact":
            vendor_id = request.POST.get('vendor_id')
            contact_id = request.POST.get('contact_id')

            vendor = Vendors.objects.get(id=vendor_id)
            contact = VendorContact.objects.get(id=contact_id, company=vendor)

            with transaction.atomic():
                clear_vendor_contact_foreign_keys(contact)
                contact.delete()

            messages.success(request, "Vendor contact was deleted.")

            return redirect(
                f"{request.path}?category_id={vendor.category.id}&vendor_id={vendor.id}"
            )

        # ------------------------------------------------------------
        # MERGE VENDOR CONTACTS
        # ------------------------------------------------------------
        elif action == "merge_vendor_contacts":
            vendor_id = request.POST.get('vendor_id')
            old_contact_id = request.POST.get('old_contact_id')
            new_contact_id = request.POST.get('new_contact_id')

            vendor = Vendors.objects.get(id=vendor_id)

            if not old_contact_id or not new_contact_id:
                messages.error(request, "Please select both contacts.")
                return redirect(
                    f"{request.path}?category_id={vendor.category.id}&vendor_id={vendor.id}"
                )

            if old_contact_id == new_contact_id:
                messages.error(request, "You cannot merge a contact into itself.")
                return redirect(
                    f"{request.path}?category_id={vendor.category.id}&vendor_id={vendor.id}"
                )

            old_contact = VendorContact.objects.get(id=old_contact_id, company=vendor)
            new_contact = VendorContact.objects.get(id=new_contact_id, company=vendor)

            conflicts = vendor_contact_conflicts(old_contact, new_contact)

            if request.POST.get('confirm_merge') != "YES" and conflicts:
                send_data['old_contact_reference_counts'] = get_vendor_contact_reference_counts(old_contact)
                send_data['contact_merge_conflicts'] = conflicts
                send_data['merge_old_contact'] = old_contact
                send_data['merge_new_contact'] = new_contact
                return render(request, 'vendor_management.html', send_data)

            old_contact_name = old_contact.name
            new_contact_name = new_contact.name
            with transaction.atomic():
                reassign_vendor_contact_foreign_keys(old_contact, new_contact)
                old_contact.delete()

            messages.success(
                request,
                f"{old_contact_name} was merged into {new_contact_name}."
            )

            return redirect(
                f"{request.path}?category_id={vendor.category.id}&vendor_id={vendor.id}"
            )

    return render(request, 'vendor_management.html', send_data)

def vendor_conflicts(old_vendor, new_vendor):
    conflicts = []

    if (old_vendor.company_name or "") != (new_vendor.company_name or ""):
        conflicts.append({
            "field": "Company Name",
            "old_value": old_vendor.company_name,
            "new_value": new_vendor.company_name,
        })

    if (old_vendor.company_phone or "") != (new_vendor.company_phone or ""):
        conflicts.append({
            "field": "Phone",
            "old_value": old_vendor.company_phone,
            "new_value": new_vendor.company_phone,
        })

    if (old_vendor.company_email or "") != (new_vendor.company_email or ""):
        conflicts.append({
            "field": "Email",
            "old_value": old_vendor.company_email,
            "new_value": new_vendor.company_email,
        })

    if old_vendor.category != new_vendor.category:
        conflicts.append({
            "field": "Category",
            "old_value": old_vendor.category,
            "new_value": new_vendor.category,
        })

    return conflicts


def vendor_contact_conflicts(old_contact, new_contact):
    conflicts = []

    if (old_contact.name or "") != (new_contact.name or ""):
        conflicts.append({
            "field": "Name",
            "old_value": old_contact.name,
            "new_value": new_contact.name,
        })

    if (old_contact.phone or "") != (new_contact.phone or ""):
        conflicts.append({
            "field": "Phone",
            "old_value": old_contact.phone,
            "new_value": new_contact.phone,
        })

    if (old_contact.email or "") != (new_contact.email or ""):
        conflicts.append({
            "field": "Email",
            "old_value": old_contact.email,
            "new_value": new_contact.email,
        })

    return conflicts

def client_cleanup_report(request):
    send_data = {}

    def normalize_text(value):
        if not value:
            return ""

        value = value.lower().strip()

        replacements = {
            "&": "and",
            ".": "",
            ",": "",
            ";": "",
            "'": "",
            '"': "",
            " inc": "",
            " incorporated": "",
            " llc": "",
            " ltd": "",
            " co": "",
            " company": "",
            " corporation": "",
            " corp": "",
            " the ": " ",
        }

        for old, new in replacements.items():
            value = value.replace(old, new)

        value = " ".join(value.split())
        return value

    def similarity(a, b):
        return SequenceMatcher(None, a, b).ratio()

    # --------------------------------------------------
    # 1. Duplicate client company names only
    # --------------------------------------------------

    clients = list(
        Clients.objects
        .filter(company__isnull=False)
        .exclude(company="")
        .order_by("company")
    )

    duplicate_clients = []
    normalized_client_map = {}

    for client in clients:
        normalized_name = normalize_text(client.company)

        if normalized_name in normalized_client_map:
            duplicate_clients.append({
                "client_1": normalized_client_map[normalized_name],
                "client_2": client,
                "reason": "Normalized company name matches exactly",
            })
        else:
            normalized_client_map[normalized_name] = client

    # --------------------------------------------------
    # 2. Duplicate / similar ClientEmployees
    # --------------------------------------------------

    client_employees = list(
        ClientEmployees.objects
        .select_related("id")
        .filter(name__isnull=False)
        .exclude(name="")
        .order_by("id__company", "name")
    )

    duplicate_client_employees = []
    similar_client_employees = []

    normalized_employee_map = {}

    for employee in client_employees:
        normalized_name = normalize_text(employee.name)
        key = f"{employee.id_id}|{normalized_name}"

        if key in normalized_employee_map:
            duplicate_client_employees.append({
                "employee_1": normalized_employee_map[key],
                "employee_2": employee,
                "reason": "Same client and same normalized employee name",
            })
        else:
            normalized_employee_map[key] = employee

    employees_by_client = {}

    for employee in client_employees:
        employees_by_client.setdefault(employee.id_id, []).append(employee)

    for client_id, employees in employees_by_client.items():
        for i in range(len(employees)):
            employee_1 = employees[i]
            employee_1_normalized = normalize_text(employee_1.name)

            for j in range(i + 1, len(employees)):
                employee_2 = employees[j]
                employee_2_normalized = normalize_text(employee_2.name)

                if not employee_1_normalized or not employee_2_normalized:
                    continue

                if employee_1_normalized == employee_2_normalized:
                    continue

                score = similarity(employee_1_normalized, employee_2_normalized)

                if score >= 0.88:
                    similar_client_employees.append({
                        "employee_1": employee_1,
                        "employee_2": employee_2,
                        "score": round(score * 100, 1),
                    })

    # --------------------------------------------------
    # 3. Bad employee names
    # --------------------------------------------------

    bad_name_employees = []

    good_name_regex = re.compile(r"^[A-Za-z\s'.-]+$")
    multiple_person_characters = [";", ",", "/", "&", " and "]

    for employee in client_employees:
        name = employee.name.strip()
        name_lower = f" {name.lower()} "

        reasons = []

        for bad_character in multiple_person_characters:
            if bad_character in name_lower:
                reasons.append(f"Contains '{bad_character.strip()}'")

        if not good_name_regex.match(name):
            reasons.append("Contains characters other than letters, spaces, apostrophes, hyphens, or periods")

        if reasons:
            bad_name_employees.append({
                "employee": employee,
                "reasons": reasons,
            })

    send_data["duplicate_clients"] = duplicate_clients
    send_data["duplicate_client_employees"] = duplicate_client_employees
    send_data["similar_client_employees"] = similar_client_employees
    send_data["bad_name_employees"] = bad_name_employees

    send_data["duplicate_clients_count"] = len(duplicate_clients)
    send_data["duplicate_client_employees_count"] = len(duplicate_client_employees)
    send_data["similar_client_employees_count"] = len(similar_client_employees)
    send_data["bad_name_employees_count"] = len(bad_name_employees)

    return render(request, "client_cleanup_report.html", send_data)

def client_merge_review(request, client_1_id, client_2_id):
    client_1 = get_object_or_404(Clients, id=client_1_id)
    client_2 = get_object_or_404(Clients, id=client_2_id)

    client_1_employees = ClientEmployees.objects.filter(id=client_1).order_by("name")
    client_2_employees = ClientEmployees.objects.filter(id=client_2).order_by("name")

    if request.method == "POST":
        keep_client_id = request.POST.get("keep_client_id")

        if str(client_1.id) == str(keep_client_id):
            keep_client = client_1
            merge_client = client_2
        else:
            keep_client = client_2
            merge_client = client_1

        with transaction.atomic():
            # Move active client employees from the duplicate company to the company being kept
            for employee in ClientEmployees.objects.filter(id=merge_client, is_active=True).order_by("name"):
                employee.id = keep_client
                employee.save()

            # Move jobs from the duplicate company to the company being kept
            for job in Jobs.objects.filter(client=merge_client):
                job.client = keep_client
                job.save()

            # Delete the duplicate company
            merged_company_name = merge_client.company
            kept_company_name = keep_client.company

            merge_client.delete()

        messages.success(
            request,
            f"{merged_company_name} was merged into {kept_company_name}."
        )

        return redirect("client_cleanup_report")

    return render(request, "client_merge_review.html", {
        "client_1": client_1,
        "client_2": client_2,
        "client_1_employees": client_1_employees,
        "client_2_employees": client_2_employees,
    })

def client_employee_merge_review(request, employee_1_id, employee_2_id):
    employee_1 = get_object_or_404(
        ClientEmployees.objects.select_related("id"),
        person_pk=employee_1_id
    )
    employee_2 = get_object_or_404(
        ClientEmployees.objects.select_related("id"),
        person_pk=employee_2_id
    )

    def default_value(value_1, value_2):
        value_1 = value_1 or ""
        value_2 = value_2 or ""

        if value_1 and not value_2:
            return value_1

        if value_2 and not value_1:
            return value_2

        return value_1

    if request.method == "POST":
        keep_employee_id = request.POST.get("keep_employee_id")

        if str(employee_1.person_pk) == str(keep_employee_id):
            keep_employee = employee_1
            merge_employee = employee_2
        else:
            keep_employee = employee_2
            merge_employee = employee_1

        final_name = request.POST.get("final_name", "").strip()
        final_phone = request.POST.get("final_phone", "").strip()
        final_email = request.POST.get("final_email", "").strip()
        final_title = request.POST.get("final_title", "").strip()

        with transaction.atomic():
            keep_employee.name = final_name
            keep_employee.phone = final_phone
            keep_employee.email = final_email
            keep_employee.title = final_title
            keep_employee.is_active = True
            keep_employee.save()

            Jobs.objects.filter(client_Pm=merge_employee).update(client_Pm=keep_employee)
            Jobs.objects.filter(client_Super=merge_employee).update(client_Super=keep_employee)

            ClientJobRoles.objects.filter(employee=merge_employee).update(employee=keep_employee)
            TempRecipients.objects.filter(person=merge_employee).update(person=keep_employee)
            TempRecipientsCOPList.objects.filter(person=merge_employee).update(person=keep_employee)

            merged_employee_name = merge_employee.name
            kept_employee_name = keep_employee.name

            merge_employee.delete()

        messages.success(
            request,
            f"{merged_employee_name} was merged into {kept_employee_name}."
        )

        return redirect("client_cleanup_report")

    send_data = {
        "employee_1": employee_1,
        "employee_2": employee_2,

        "default_name": default_value(employee_1.name, employee_2.name),
        "default_phone": default_value(employee_1.phone, employee_2.phone),
        "default_email": default_value(employee_1.email, employee_2.email),
        "default_title": default_value(employee_1.title, employee_2.title),
    }

    return render(request, "client_employee_merge_review.html", send_data)
