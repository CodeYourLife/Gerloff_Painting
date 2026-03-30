import os
from datetime import datetime, date

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from jobs.models import Jobs
from submittals.models import Submittals, SubmittalItems, SubmittalApprovals
from console.misc import createfolder, getFilesOrFolders, create_shortcut

import os
from datetime import datetime, date

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from jobs.models import Jobs
from employees.models import Employees
from submittals.models import (
    Submittals,
    SubmittalItems,
    SubmittalApprovals,
    SubmittalNotes,
)
from console.misc import createfolder


class Command(BaseCommand):
    help = "Import Submittals, SubmittalItems, SubmittalApprovals, and SubmittalNotes from Excel"

    def add_arguments(self, parser):
        parser.add_argument("excel_path", type=str, help="Path to the Excel file")
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Parse and validate, but do not save anything",
        )

    def handle(self, *args, **options):
        excel_path = options["excel_path"]
        dry_run = options["dry_run"]

        if not os.path.exists(excel_path):
            raise CommandError(f"File does not exist: {excel_path}")

        wb = load_workbook(excel_path, data_only=True)

        if "submittals" not in wb.sheetnames:
            raise CommandError("Excel file is missing sheet: submittals")
        if "items" not in wb.sheetnames:
            raise CommandError("Excel file is missing sheet: items")

        submittals_ws = wb["submittals"]
        items_ws = wb["items"]

        self.stdout.write(self.style.NOTICE("Reading workbook..."))

        submittal_rows = self.read_sheet(submittals_ws)
        item_rows = self.read_sheet(items_ws)

        self.stdout.write(
            self.style.NOTICE(
                f"Found {len(submittal_rows)} submittal rows and {len(item_rows)} item rows"
            )
        )

        try:
            with transaction.atomic():
                results = self.import_data(
                    submittal_rows=submittal_rows,
                    item_rows=item_rows,
                    dry_run=dry_run,
                )

                if dry_run:
                    self.stdout.write(self.style.WARNING("Dry run requested. Rolling back."))
                    raise DryRunRollback()

        except DryRunRollback:
            self.stdout.write(self.style.SUCCESS("Dry run completed successfully."))
            return

        self.stdout.write(self.style.SUCCESS("Import completed."))
        self.stdout.write(f"Submittals created: {results['submittals_created']}")
        self.stdout.write(f"Submittals reused:  {results['submittals_reused']}")
        self.stdout.write(f"Items created:      {results['items_created']}")
        self.stdout.write(f"Approvals created:  {results['approvals_created']}")
        self.stdout.write(f"Approvals reused:   {results['approvals_reused']}")
        self.stdout.write(f"Notes created:      {results['notes_created']}")
        self.stdout.write(f"Notes reused:       {results['notes_reused']}")
        self.stdout.write(f"Rows skipped:       {results['rows_skipped']}")

        if results["missing_jobs"]:
            self.stdout.write(self.style.WARNING("Missing Jobs:"))
            for job_number in sorted(set(results["missing_jobs"])):
                self.stdout.write(f" - {job_number}")

    def read_sheet(self, ws):
        headers = [self.clean_header(cell.value) for cell in ws[1]]

        rows = []
        for row_idx in range(2, ws.max_row + 1):
            values = [
                ws.cell(row=row_idx, column=col_idx).value
                for col_idx in range(1, len(headers) + 1)
            ]

            if self.is_blank_row(values):
                continue

            row_data = {}
            for header, value in zip(headers, values):
                row_data[header] = value

            row_data["_excel_row"] = row_idx
            rows.append(row_data)

        return rows

    def import_data(self, submittal_rows, item_rows, dry_run=False):
        results = {
            "submittals_created": 0,
            "submittals_reused": 0,
            "items_created": 0,
            "approvals_created": 0,
            "approvals_reused": 0,
            "notes_created": 0,
            "notes_reused": 0,
            "rows_skipped": 0,
            "missing_jobs": [],
        }

        note_user = Employees.objects.get(id=42)
        submittal_lookup = {}

        # ------------------------------------------------------------------
        # 1. CREATE / REUSE SUBMITTALS + NOTES
        # ------------------------------------------------------------------
        for row in submittal_rows:
            excel_row = row["_excel_row"]

            job_number_value = self.clean_string(row.get("job_number"))
            submittal_number_value = self.to_int(row.get("submittal_number"))
            description_value = self.clean_string(row.get("description"))
            date_sent_value = self.to_date(row.get("date_sent"))
            originated_value = self.to_bool(row.get("originated_in_management_console"))
            notes_value = self.clean_string(row.get("notes"))

            if description_value and len(description_value) > 2000:
                self.stdout.write(
                    self.style.WARNING(
                        f"Submittals row {excel_row}: description longer than 2000 chars. Truncating."
                    )
                )
                description_value = description_value[:2000]

            if not job_number_value or submittal_number_value is None:
                self.stdout.write(
                    self.style.WARNING(
                        f"Skipping submittals row {excel_row}: missing job_number or submittal_number"
                    )
                )
                results["rows_skipped"] += 1
                continue

            job_obj = self.get_job(job_number_value)
            if job_obj is None:
                self.stdout.write(
                    self.style.WARNING(
                        f"Skipping submittals row {excel_row}: could not find Jobs object with job_number={job_number_value}"
                    )
                )
                results["rows_skipped"] += 1
                results["missing_jobs"].append(job_number_value)
                continue

            self.stdout.write(
                f"Row {excel_row} | Job {job_number_value} | Sub {submittal_number_value} | "
                f"NOTES FROM EXCEL: {notes_value[:80] if notes_value else 'None'}"
            )

            submittal_obj, created = Submittals.objects.get_or_create(
                job_number=job_obj,
                submittal_number=submittal_number_value,
                defaults={
                    "description": description_value,
                    "date_sent": date_sent_value,
                    "originated_in_management_console": (
                        originated_value if originated_value is not None else True
                    ),
                },
            )

            updated = False

            if description_value is not None and description_value != submittal_obj.description:
                submittal_obj.description = description_value
                updated = True

            if date_sent_value is not None and date_sent_value != submittal_obj.date_sent:
                submittal_obj.date_sent = date_sent_value
                updated = True

            if (
                originated_value is not None
                and originated_value != submittal_obj.originated_in_management_console
            ):
                submittal_obj.originated_in_management_console = originated_value
                updated = True

            if created:
                results["submittals_created"] += 1

                if not dry_run:
                    try:
                        folder_name = (
                            f"{submittal_obj.job_number.job_number} "
                            f"{submittal_obj.submittal_number}"
                        )
                        base_path = f"submittals/{folder_name}"

                        createfolder(base_path)
                        createfolder(f"{base_path}/Sent to GC")
                        createfolder(f"{base_path}/Approval Documents")

                        self.stdout.write(f"📁 Created folders for {folder_name}")

                    except Exception as e:
                        self.stdout.write(
                            self.style.WARNING(
                                f"⚠️ Failed to create folders for Submittal ID {submittal_obj.id}: {str(e)}"
                            )
                        )
            else:
                results["submittals_reused"] += 1

            if updated:
                submittal_obj.save()
                self.stdout.write(
                    f"✅ SAVED Submittal {job_number_value}-{submittal_number_value}"
                )

            # --------------------------------------------------------------
            # CREATE SUBMITTAL NOTE FROM EXCEL NOTES
            # --------------------------------------------------------------
            if notes_value:
                existing_note = SubmittalNotes.objects.filter(
                    submittal=submittal_obj,
                    note=notes_value,
                ).first()

                if existing_note:
                    results["notes_reused"] += 1
                    self.stdout.write(
                        f"🔁 Note already exists for {job_number_value}-{submittal_number_value}"
                    )
                else:
                    SubmittalNotes.objects.create(
                        submittal=submittal_obj,
                        date=date.today(),
                        user=note_user,
                        note=notes_value,
                    )
                    results["notes_created"] += 1
                    self.stdout.write(
                        f"📝 Added SubmittalNote | {job_number_value}-{submittal_number_value} | "
                        f"{notes_value[:80]}"
                    )

            submittal_lookup[(job_number_value, submittal_number_value)] = submittal_obj

        # ------------------------------------------------------------------
        # 2. CREATE ITEMS + APPROVALS
        # ------------------------------------------------------------------
        for row in item_rows:
            excel_row = row["_excel_row"]

            job_number_value = self.clean_string(row.get("job_number"))
            item_description_value = self.clean_string(row.get("description"))
            quantity_value = self.to_int(row.get("copies"), default=0)
            date_reviewed_value = self.to_date(row.get("date_reviewed"))
            submittal_number_value = self.to_int(row.get("submittal"))
            is_approved_value = self.to_bool(row.get("is_approved"))

            if not job_number_value or not item_description_value:
                self.stdout.write(
                    self.style.WARNING(
                        f"Skipping items row {excel_row}: missing job_number or description"
                    )
                )
                results["rows_skipped"] += 1
                continue

            if len(item_description_value) > 250:
                self.stdout.write(
                    self.style.WARNING(
                        f"Items row {excel_row}: description longer than 250 chars. Truncating."
                    )
                )
                item_description_value = item_description_value[:250]

            job_obj = self.get_job(job_number_value)
            if job_obj is None:
                self.stdout.write(
                    self.style.WARNING(
                        f"Skipping items row {excel_row}: could not find Jobs object with job_number={job_number_value}"
                    )
                )
                results["rows_skipped"] += 1
                results["missing_jobs"].append(job_number_value)
                continue

            submittal_obj = None
            if submittal_number_value is not None:
                submittal_obj = submittal_lookup.get((job_number_value, submittal_number_value))
                if submittal_obj is None:
                    submittal_obj = Submittals.objects.filter(
                        job_number=job_obj,
                        submittal_number=submittal_number_value,
                    ).first()

            if submittal_number_value is not None and submittal_obj is None:
                self.stdout.write(
                    self.style.WARNING(
                        f"Skipping items row {excel_row}: could not find Submittal for "
                        f"job_number={job_number_value}, submittal_number={submittal_number_value}"
                    )
                )
                results["rows_skipped"] += 1
                continue

            item_obj = self.find_existing_item(
                job_obj=job_obj,
                description=item_description_value,
            )

            if item_obj is None:
                item_obj = SubmittalItems.objects.create(
                    wallcovering_id=None,
                    description=item_description_value,
                    notes="",
                    is_no_longer_used=False,
                    job_number=job_obj,
                )
                results["items_created"] += 1

            approval_qs = SubmittalApprovals.objects.filter(
                submittal=submittal_obj,
                submittalitem=item_obj,
            )

            if approval_qs.exists():
                approval_obj = approval_qs.first()
                updated = False
                changes = []

                if approval_obj.quantity != quantity_value:
                    changes.append(f"quantity {approval_obj.quantity} → {quantity_value}")
                    approval_obj.quantity = quantity_value
                    updated = True

                if approval_obj.date_reviewed is None and date_reviewed_value is not None:
                    changes.append(f"date_reviewed None → {date_reviewed_value}")
                    approval_obj.date_reviewed = date_reviewed_value
                    updated = True

                if approval_obj.is_approved is None and is_approved_value is not None:
                    changes.append(f"is_approved None → {is_approved_value}")
                    approval_obj.is_approved = is_approved_value
                    updated = True

                if updated:
                    approval_obj.save()

                self.stdout.write(
                    f"🔁 Reused Approval | Job: {job_number_value} | "
                    f"Submittal: {submittal_number_value} | "
                    f"Item: {item_description_value[:50]} | "
                    f"Changes: {', '.join(changes) if changes else 'No changes'}"
                )

                results["approvals_reused"] += 1
            else:
                SubmittalApprovals.objects.create(
                    submittal=submittal_obj,
                    submittalitem=item_obj,
                    is_approved=is_approved_value,
                    notes="",
                    quantity=quantity_value or 0,
                    date_reviewed=date_reviewed_value,
                )
                results["approvals_created"] += 1

        return results

    def get_job(self, job_number_value):
        cleaned_job_number = (job_number_value or "").strip()
        return Jobs.objects.filter(job_number=cleaned_job_number).first()

    def find_existing_item(self, job_obj, description):
        return SubmittalItems.objects.filter(
            job_number=job_obj,
            description=description,
        ).first()

    def clean_header(self, value):
        if value is None:
            return ""
        return str(value).strip().lower()

    def clean_string(self, value):
        if value is None:
            return None
        value = str(value).strip()
        return value if value != "" else None

    def to_int(self, value, default=None):
        if value in (None, ""):
            return default
        try:
            return int(value)
        except (TypeError, ValueError):
            return default

    def to_date(self, value):
        if value in (None, ""):
            return None

        if isinstance(value, datetime):
            return value.date()

        if isinstance(value, date):
            return value

        if isinstance(value, str):
            value = value.strip()
            if not value:
                return None

            for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    pass

        return None

    def to_bool(self, value):
        if value in (None, ""):
            return None

        if isinstance(value, bool):
            return value

        if isinstance(value, (int, float)):
            return bool(value)

        if isinstance(value, str):
            v = value.strip().lower()
            if v in ("true", "yes", "y", "1"):
                return True
            if v in ("false", "no", "n", "0"):
                return False

        return None

    def is_blank_row(self, values):
        return all(v in (None, "") for v in values)


class DryRunRollback(Exception):
    pass